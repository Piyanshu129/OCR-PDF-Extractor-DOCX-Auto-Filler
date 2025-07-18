import streamlit as st
import pandas as pd
import io
import xlwt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime
#pip uninstall numpy pandas -y
#pip install numpy pandas
# === Constants ===

#paths line 14,15,16,755,769,1031,1539,1607,1644,2013,2027,2441,3113,3133,3170,3761,3798
COUNTRY_CODE_PATH = "/Users/piyanshu/PycharmProjects/pdftoexcel/Country Code.xlsx"  # 👈 change if needed
UQC_MAPPING_PATH = "/Users/piyanshu/PycharmProjects/pdftoexcel/SQC_1.xlsx"          # 👈 change if needed
SAVE_DIRECTORY = "/Users/piyanshu/Downloads"  # 👈 your preferred save location

# === App Config ===
# === App Config ===
st.set_page_config(layout="wide")
st.title("🧾 GST Invoice to Formatted Data Extractor")

# Dropdown selection for company
company_options = ["Bhajan", "Gupta International","PACHRANGA FOODS","Imperial","Bhajan-sqm","Javi","Tax India","Tax India - d","Tax India - d1"]
selected_company = st.selectbox("Select Company", company_options)

if selected_company == "Gupta International":
    uploaded_file = st.file_uploader("Upload GST Invoice Excel File", type=["xlsx"])
    dollar_price = st.number_input("💵 Enter Dollar Price (Exchange Rate)", min_value=0.0, value=83.0, step=0.1)
    st.info(f"🧮 TAXABLE_VALUE will be calculated as TOTAL_VAL_FC × {dollar_price}")

    if uploaded_file:
        # Existing code for Gupta International
        df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)

        # Extract company name and invoice number
        company_name = ""
        invoice_number = ""

        for row in df_raw[0].dropna().head(20):
            row_str = str(row).strip()
            if not company_name and any(
                    x in row_str.upper() for x in ["INTERNATIONAL", "LIMITED", "LTD", "PVT", "LLP"]):
                company_name = row_str.split(",")[0].strip()
            if "invoice no" in row_str.lower():
                try:
                    invoice_number = row_str.split("/")[-1].strip()
                except:
                    invoice_number = "XXXX"

        clean_co_name = company_name.upper().replace(" ", "_").replace(",", "")
        file_name = f"{clean_co_name}_{invoice_number}.xlsx"

    # === Load Country Mapping ===
    try:
        country_df = pd.read_excel(COUNTRY_CODE_PATH)
        country_df['Country'] = country_df['Country'].astype(str).str.strip().str.upper()
        country_to_code = dict(zip(country_df['Country'], country_df['Code']))
        country_to_code['ANYTHING ELSE'] = 'NCPTI'
    except Exception as e:
        st.error(f"❌ Failed to load country code mapping: {e}")
        country_to_code = {}

    if uploaded_file:
        df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
        # st.subheader("Raw Excel Preview")
        # st.dataframe(df_raw.head(25))

        # === Extract Country ===
        detected_country = None
        for i in range(min(40, len(df_raw))):
            row = df_raw.iloc[i]
            for j in range(len(row) - 1):
                if isinstance(row[j], str) and 'country' in row[j].lower():
                    detected_country = str(row[j + 1]).strip().upper()
                    break
            if detected_country:
                break
        st.success(f"🌍 Detected Country: {detected_country if detected_country else 'NOT FOUND'}")

        # === Get SWI_EPT Code ===
        swi_ept_code = country_to_code.get(detected_country, country_to_code.get('ANYTHING ELSE', 'NCPTI'))
        swi_ept_code = st.text_input("📄 SWI_EPT Code (editable)", value=swi_ept_code)

        # === Header Detection ===
        header_keywords = ["Sr", "Description", "HSN", "Qty", "Rate", "Amount", "Taxable", "IGST", "Total"]
        header_row_index = None
        for i, row in df_raw.iterrows():
            match_count = sum(
                any(k.lower() in str(cell).lower() for cell in row if pd.notna(cell)) for k in header_keywords)
            if match_count >= 4:
                header_row_index = i
                break

        if header_row_index is not None:
            df = pd.read_excel(uploaded_file, sheet_name=0, header=header_row_index)
            df = df.dropna(how='all')
            df.columns = [str(c).strip() for c in df.columns]
            st.subheader("Detected Invoice Table")
            st.dataframe(df)

            # === Enhanced Column Mapping with USD/GBP condition ===
            # Base column mapping
            column_mapping = {
                "Sr. No.": "ITEM_SR_NO", "Description": "GOODS_DESC1", "Product": "GOODS_DESC1",
                "HSN": "RITC", "Code": "RITC", "Qty.": "QTY_NOS", "Unnamed: 4": "QTY_NOS",
                "Taxable value.1": "TAXABLE_VALUE", "IGST": "IGST_RATE",
                "In Set": "UNIT_QTY"
            }

            # Check for USD or GBP columns and map accordingly
            available_columns = df.columns.tolist()

            # For RATE_VALUE - check for "in USD" or "in GBP"
            if "in USD" in available_columns:
                column_mapping["in USD"] = "RATE_VALUE"
                st.info("💰 Detected currency: USD")
            elif "in GBP" in available_columns:
                column_mapping["in GBP"] = "RATE_VALUE"
                st.info("💰 Detected currency: GBP")
            else:
                st.warning("⚠️ No USD or GBP rate column found")

            # For TOTAL_VAL_FC - check for "in USD.1" or "in GBP.1"
            if "in USD.1" in available_columns:
                column_mapping["in USD.1"] = "TOTAL_VAL_FC"
            elif "in GBP.1" in available_columns:
                column_mapping["in GBP.1"] = "TOTAL_VAL_FC"
            elif "in USD" in available_columns and "in USD" not in column_mapping:
                # If there's only one USD column, use it for TOTAL_VAL_FC
                column_mapping["in USD"] = "TOTAL_VAL_FC"
            elif "in GBP" in available_columns and "in GBP" not in column_mapping:
                # If there's only one GBP column, use it for TOTAL_VAL_FC
                column_mapping["in GBP"] = "TOTAL_VAL_FC"

            mapped_df = pd.DataFrame()
            for col, new_col in column_mapping.items():
                if col in df.columns:
                    mapped_df[new_col] = df[col]


            # === Standardize UNIT_QTY values ===
            def standardize_unit_qty(value):
                if isinstance(value, str) and "pcs" in value.lower():
                    return "PCS"
                return value


            if 'UNIT_QTY' in mapped_df.columns:
                mapped_df['UNIT_QTY'] = mapped_df['UNIT_QTY'].apply(standardize_unit_qty)

            # === Footer Cleanup ===
            footer_keywords = ["Total Invoice amount", "Bank Name", "Bank A/C", "Certified", "Total Amount",
                               "GST ON Reverse"]


            def is_footer_row(row):
                return any(any(kw.lower() in str(cell).lower() for kw in footer_keywords) for cell in row)


            mapped_df = mapped_df[~mapped_df.apply(is_footer_row, axis=1)].reset_index(drop=True)

            # === Enhanced Filtering - Remove empty rows ===
            # Remove rows where all key columns are empty or contain only whitespace
            key_columns = ["ITEM_SR_NO", "RITC", "GOODS_DESC1", "QTY_NOS", "TOTAL_VAL_FC"]


            def is_empty_row(row):
                for col in key_columns:
                    if col in row.index:
                        value = row[col]
                        if pd.notna(value) and str(value).strip() not in ["", "0", "nan", "0.0"]:
                            return False
                return True


            # Remove completely empty rows
            mapped_df = mapped_df[~mapped_df.apply(is_empty_row, axis=1)].reset_index(drop=True)

            # Additional filtering for specific columns
            for field in ["ITEM_SR_NO", "RITC"]:
                if field in mapped_df.columns:
                    mapped_df = mapped_df[
                        mapped_df[field].notna() &
                        ~(mapped_df[field].astype(str).str.strip().isin(["", "0", "nan", "0.0"]))
                        ]

            # Remove rows where GOODS_DESC1 is empty
            if "GOODS_DESC1" in mapped_df.columns:
                mapped_df = mapped_df[
                    mapped_df["GOODS_DESC1"].notna() &
                    (mapped_df["GOODS_DESC1"].astype(str).str.strip() != "")
                    ]

            # Reset index after all filtering
            mapped_df = mapped_df.reset_index(drop=True)

            # === Calculate TAXABLE_VALUE and IGST_AMOUNT ===
            if "TOTAL_VAL_FC" in mapped_df.columns:
                mapped_df["TAXABLE_VALUE"] = pd.to_numeric(mapped_df["TOTAL_VAL_FC"], errors='coerce') * dollar_price

            if "IGST_RATE" in mapped_df.columns and "TAXABLE_VALUE" in mapped_df.columns:
                mapped_df["IGST_AMOUNT"] = pd.to_numeric(mapped_df["IGST_RATE"], errors='coerce') * pd.to_numeric(
                    mapped_df["TAXABLE_VALUE"], errors='coerce')

            # === Output ===
            st.subheader("📄 Cleaned Mapped Table with IGST Amount")
            st.dataframe(mapped_df)

            # === Final Format ===
            required_columns = [
                "INVOICE_SR_NO", "ITEM_SR_NO", "SCHEME_CODE", "RITC", "GOODS_DESC1", "GOODS_DESC2", "GOODS_DESC3",
                "QTY_NOS", "UNIT_QTY", "RATE_VALUE", "NO_OF_UNIT", "UNIT_OF_RATE", "PMV_AMT", "TOTAL_PMV",
                "ACCESSORIES_FLG", "CESS_FLG", "THIRD_PARTY_FLG", "AR4_FLG", "REWARD_FLG", "TOTAL_VAL_FC",
                "STR_FLG", "END_USE", "IGST_PAYMENT_STATUS", "TAXABLE_VALUE", "IGST_AMOUNT", "SWI_STO",
                "SWI_DOO", "SWI_EPT", "SWI_UQC", "SWI_QTY", "SWI_GCESS_AMT", "SWI_GCESS_CUR", "RODTEP_FLG",
                "SOURCE_STATE", "DBK_SRNO", "DBK_QUANTITY"
            ]

            output_df = pd.DataFrame(columns=required_columns)

            # Only copy data if mapped_df has rows
            if not mapped_df.empty:
                for col in output_df.columns:
                    if col in mapped_df.columns:
                        output_df[col] = mapped_df[col]

                # Set default values
                output_df["INVOICE_SR_NO"] = 1
                output_df["GOODS_DESC2"] = "SIZE:"
                output_df["NO_OF_UNIT"] = 1
                output_df["UNIT_OF_RATE"] = output_df["UNIT_QTY"]
                output_df["ACCESSORIES_FLG"] = "N"
                output_df["CESS_FLG"] = "N"
                output_df["THIRD_PARTY_FLG"] = "N"
                output_df["STR_FLG"] = "N"
                output_df["AR4_FLG"] = "N"
                output_df["REWARD_FLG"] = "Y"
                output_df["IGST_PAYMENT_STATUS"] = "P"
                output_df["SWI_STO"] = "06"
                output_df["SWI_DOO"] = "71"
                output_df["SWI_EPT"] = swi_ept_code
                output_df["SWI_GCESS_CUR"] = "INR"
                output_df["SOURCE_STATE"] = "06"
                output_df["SWI_GCESS_AMT"] = "0"

                # SCHEME & RODTEP
                output_df["SCHEME_CODE"] = output_df["RITC"].astype(str).apply(
                    lambda x: "60" if x.startswith(("60", "61", "62", "63")) else "19")
                output_df["RODTEP_FLG"] = output_df["SCHEME_CODE"].apply(lambda x: "N" if x == "60" else "Y")

                # === Load UQC Mapping ===
                try:
                    uqc_map_df = pd.read_excel(UQC_MAPPING_PATH)
                    uqc_map_df['RITC'] = uqc_map_df['RITC'].astype(str).str.strip()
                    ritc_to_uqc = dict(zip(uqc_map_df['RITC'], uqc_map_df['SWI_UQC']))
                    output_df['SWI_UQC'] = output_df['RITC'].astype(str).str.strip().map(ritc_to_uqc).fillna("NOS")
                except Exception as e:
                    st.error(f"⚠️ Failed to load UQC mapping: {e}")
                    output_df['SWI_UQC'] = "NOS"

                # === Final check to ensure no empty rows in output ===
                output_df = output_df[output_df["ITEM_SR_NO"].notna() & (output_df["ITEM_SR_NO"] != "")].reset_index(
                    drop=True)

                # === Excel Output ===
                if not output_df.empty:
                    output = io.BytesIO()

                    # Create workbook and worksheet using xlwt (for Excel 97-2003 format)
                    workbook = xlwt.Workbook()
                    worksheet = workbook.add_sheet("FormattedInvoice")

                    # Write headers
                    for col_idx, col_name in enumerate(output_df.columns):
                        worksheet.write(0, col_idx, str(col_name))

                    # Write data
                    for row_idx, row in output_df.iterrows():
                        for col_idx, value in enumerate(row):
                            # Handle different types of data
                            if pd.isna(value):
                                worksheet.write(row_idx + 1, col_idx, '')
                            else:
                                worksheet.write(row_idx + 1, col_idx, value)

                    # Save to buffer
                    workbook.save(output)
                    output.seek(0)

                    # Save and Download
                    if st.button("📁 Add to Folder"):
                        timestamp = datetime.now().strftime("%d%m%Y_%H%M%S")
                        save_path = os.path.join(SAVE_DIRECTORY, f"Gupta_invoice_{timestamp}.xls")
                        try:
                            with open(save_path, "wb") as f:
                                f.write(output.getbuffer())
                            st.success(f"✅ File saved to: {save_path}")
                        except Exception as e:
                            st.error(f"❌ Failed to save file: {e}")

                    st.download_button(
                        "📥 Download Formatted Excel (97-2003 Format)",
                        output,
                        file_name if "file_name" in locals() else "formatted_invoice.xls",
                        mime="application/vnd.ms-excel"
                    )
                else:
                    st.warning("⚠️ No valid data rows found after processing.")
            else:
                st.warning("⚠️ No valid")




####
#  Bhajana start



def process_bhajan_to_export_format(df):
    """
    Process Bhajan invoices to match the to_export.csv format.
    Creates separate rows for sub-items with different rates.

    Args:
        df (pd.DataFrame): The DataFrame after initial header detection and cleaning.

    Returns:
        pd.DataFrame: A simplified DataFrame matching to_export format.
    """
    import pandas as pd

    # Make a copy to work on
    df_copy = df.copy()

    # --- 1. Initial Cleaning: Remove footers and non-item rows ---
    df_copy['Sr. No.'] = pd.to_numeric(df_copy['Sr. No.'], errors='coerce')

    # --- 2. Grouping Logic: Identify all rows belonging to one original item ---
    df_copy['Group_ID'] = df_copy['Sr. No.'].ffill()
    df_copy.dropna(subset=['Group_ID'], inplace=True)

    # --- 3. Process Groups and Create Records for Each Sub-item with Rate ---
    final_records = []
    output_sr_no = 1  # Counter for the output Sr. No.

    for _, group in df_copy.groupby('Group_ID'):
        # The first row of any group is the "parent" item
        parent_row = group.iloc[0]
        parent_description = str(parent_row['Product']).strip() if pd.notna(parent_row['Product']) else ""

        # Find all sub-items that have both Rate and 'in USD' values
        sub_items_with_rates = []

        for i in range(1, len(group)):
            sub_row = group.iloc[i]

            # Skip if no product description
            if pd.isna(sub_row['Product']) or not str(sub_row['Product']).strip():
                continue

            # Check if this sub-item has rate and USD values
            if pd.notna(sub_row.get('Rate')) and pd.notna(sub_row.get('in USD')):
                sub_items_with_rates.append({
                    'description': str(sub_row['Product']).strip(),
                    'rate': sub_row.get('Rate'),
                    'usd': sub_row.get('in USD')
                })

        # If no sub-items with rates found, check if parent has rate
        if not sub_items_with_rates:
            if pd.notna(parent_row.get('Rate')) and pd.notna(parent_row.get('in USD')):
                sub_items_with_rates.append({
                    'description': parent_description,
                    'rate': parent_row.get('Rate'),
                    'usd': parent_row.get('in USD')
                })

        # Create output records - one for each sub-item with rate
        for sub_item in sub_items_with_rates:
            # Combine parent description with sub-item description
            if parent_description and sub_item['description']:
                if parent_description != sub_item['description']:
                    consolidated_description = f"{parent_description} / {sub_item['description']}"
                else:
                    consolidated_description = parent_description
            else:
                consolidated_description = parent_description or sub_item['description']

            record = {
                'Sr. No.': output_sr_no,
                'Product': consolidated_description,
                'Rate': sub_item['rate'],
                'in USD': sub_item['usd'],
                'Taxable value': sub_item['usd']  # Same as 'in USD'
            }

            final_records.append(record)
            output_sr_no += 1

    if not final_records:
        return pd.DataFrame()

    # Create the final DataFrame with only the required columns
    final_df = pd.DataFrame(final_records)

    # Ensure we have the exact columns as in to_export.csv
    target_columns = ['Sr. No.', 'Product', 'Rate', 'in USD', 'Taxable value']

    # Reorder columns to match target format
    final_df = final_df[target_columns]

    return final_df.reset_index(drop=True)


from decimal import Decimal, getcontext, ROUND_HALF_UP

getcontext().prec = 20  # set high precision globally

def compute_precise_rate(row):
    try:
        total = Decimal(str(row.get('in usd/  Pcs') or row.get('in USD') or 0))
        qty = Decimal(str(row.get('Qty (in pcs)') or 1))
        rate = total / qty
        return float(rate.quantize(Decimal('1.0000000000000'), rounding=ROUND_HALF_UP))  # 13 decimal places
    except Exception:
        return 0.0


import pandas as pd
import io
import xlwt


def process_bhajan_to_export_format_alternative(df, remove_after_slash=False):
    """
    Enhanced version that processes Bhajan invoices and maps additional fields like HSN, Qty, etc.
    Produces consolidated descriptions and all important numeric fields.

    Parameters:
    df - The input DataFrame
    remove_after_slash - If True, keeps only the part before the first slash in product descriptions
    """
    import pandas as pd

    df_copy = df.copy()

    # Clean and group
    df_copy['Sr. No.'] = pd.to_numeric(df_copy['Sr. No.'], errors='coerce')
    df_copy['Group_ID'] = df_copy['Sr. No.'].ffill()
    df_copy.dropna(subset=['Group_ID'], inplace=True)

    final_records = []
    output_sr_no = 1

    for _, group in df_copy.groupby('Group_ID'):
        parent_row = group.iloc[0]
        parent_description = str(parent_row.get('Product', '')).strip()

        # Gather sub-descriptions and rows with rates
        rate_rows = []
        sub_descriptions = []

        for i in range(1, len(group)):
            sub_row = group.iloc[i]

            desc = str(sub_row.get('Product', '')).strip()
            if desc:
                sub_descriptions.append(desc)

            if pd.notna(sub_row.get('Rate')) and pd.notna(sub_row.get('in USD')):
                rate_rows.append(sub_row)

        all_descriptions = [parent_description] + sub_descriptions
        consolidated_description = " / ".join(filter(None, all_descriptions))

        # Clean up product descriptions - keep only part before first slash if option is enabled
        if remove_after_slash and '/' in consolidated_description:
            # Split at the first forward slash and keep only the first part
            parts = consolidated_description.split('/', 1)
            consolidated_description = parts[0].strip()

        for rate_row in rate_rows:
            record = {
                'Sr. No.': output_sr_no,
                'Product': consolidated_description,
                'Unit_qty': rate_row.get('Unnamed: 5') or rate_row.get('In Pcs'),
                'in USD': rate_row.get('in USD'),
                'Taxable value(in USD)': rate_row.get('in USD'),  # Redundant but kept for clarity

                # Additional fields below
                'HSN Code': rate_row.get('HSN') or rate_row.get('HSN Code'),
                'Qty (in pcs)': rate_row.get('Unnamed: 4') or rate_row.get('Qty.'),
                'Qty (Sq./Mtrs)': rate_row.get('Sq.') or rate_row.get('Mtrs'),  # in usd/  Pcs
                'Exchange Rate (in usd/  Pcs)': rate_row.get('Rate') or rate_row.get('in usd/  Pcs'),
                'Taxable Value (in Rs)': rate_row.get('Taxable value.1'),
                'IGST Rate (%)': rate_row.get('IGST'),
                'IGST Amount (in Rs)': rate_row.get('Amount in Rs.'),
                'Total (in Rs)': rate_row.get('Total in Rs.')
            }

            final_records.append(record)
            output_sr_no += 1

    if not final_records:
        return pd.DataFrame()

    final_df = pd.DataFrame(final_records)

    # Final column order
    target_columns = [
        'Sr. No.', 'Product', 'Unit_qty', 'in USD', 'Taxable value(in USD)',
        'HSN Code', 'Qty (in pcs)', 'Qty (Sq./Mtrs)', 'Exchange Rate (in usd/  Pcs)',
        'Taxable Value (in Rs)', 'IGST Rate (%)', 'IGST Amount (in Rs)', 'Total (in Rs)'
    ]

    final_df = final_df[target_columns]

    return final_df.reset_index(drop=True)


def create_final_mapped_excel(export_df2, dollar_price, uqc_mapping_path=None, swi_ept_code="NCPTI"):
    # Optional UQC Mapping
    ritc_to_uqc = {}
    if uqc_mapping_path:
        try:
            uqc_map_df = pd.read_excel(uqc_mapping_path)
            uqc_map_df['RITC'] = uqc_map_df['RITC'].astype(str).str.strip()
            ritc_to_uqc = dict(zip(uqc_map_df['RITC'], uqc_map_df['SWI_UQC']))
        except Exception as e:
            st.error(f"⚠️ Failed to load UQC mapping: {e}")
            ritc_to_uqc = {}

    output_df = pd.DataFrame()
    output_df['INVOICE_SR_NO'] = [1] * len(export_df2)
    output_df['ITEM_SR_NO'] = export_df2['Sr. No.']
    output_df['SCHEME_CODE'] = export_df2['HSN Code'].astype(str).apply(
        lambda x: "60" if x.startswith(("60", "61", "62", "63")) else "19")
    output_df['RITC'] = export_df2['HSN Code']
    output_df['GOODS_DESC1'] = export_df2['Product']
    output_df['GOODS_DESC2'] = "SIZE:"
    output_df['GOODS_DESC3'] = "As Per Invoice"
    output_df['QTY_NOS'] = export_df2['Qty (in pcs)']
    output_df['UNIT_QTY'] = export_df2['Unit_qty'].astype(str).str.upper()
    output_df['RATE_VALUE'] = export_df2['Exchange Rate (in usd/  Pcs)']
    output_df['NO_OF_UNIT'] = 1
    output_df['UNIT_OF_RATE'] = output_df['UNIT_QTY']
    output_df['PMV_AMT'] = ""
    output_df['TOTAL_PMV'] = ""
    output_df['ACCESSORIES_FLG'] = "N"
    output_df['CESS_FLG'] = "N"
    output_df['THIRD_PARTY_FLG'] = "N"
    output_df['AR4_FLG'] = "N"
    output_df['REWARD_FLG'] = "Y"
    output_df['TOTAL_VAL_FC'] = export_df2['Taxable value(in USD)']
    output_df['STR_FLG'] = "N"
    output_df['END_USE'] = "GNX100"
    output_df['IGST_PAYMENT_STATUS'] = "P"
    output_df['TAXABLE_VALUE'] = pd.to_numeric(output_df["TOTAL_VAL_FC"], errors='coerce') * dollar_price
    output_df['IGST_AMOUNT'] = (pd.to_numeric(export_df2["IGST Rate (%)"], errors='coerce')) * output_df[
        "TAXABLE_VALUE"]
    output_df['SWI_STO'] = "06"
    output_df['SWI_DOO'] = "71"
    output_df['SWI_EPT'] = swi_ept_code
    output_df['SWI_UQC'] = output_df['RITC'].astype(str).str.strip().map(ritc_to_uqc).fillna("NOS")
    output_df['SWI_QTY'] = ""
    output_df['SWI_GCESS_AMT'] = 0
    output_df['SWI_GCESS_CUR'] = "INR"
    output_df['RODTEP_FLG'] = output_df['SCHEME_CODE'].apply(lambda x: "N" if x == "60" else "Y")
    output_df['SOURCE_STATE'] = "06"
    output_df['DBK_SRNO'] = ""
    output_df['DBK_QUANTITY'] = ""

    # Final Column Order as per your specification
    final_columns = [
        'INVOICE_SR_NO', 'ITEM_SR_NO', 'SCHEME_CODE', 'RITC', 'GOODS_DESC1', 'GOODS_DESC2', 'GOODS_DESC3',
        'QTY_NOS', 'UNIT_QTY', 'RATE_VALUE', 'NO_OF_UNIT', 'UNIT_OF_RATE', 'PMV_AMT', 'TOTAL_PMV',
        'ACCESSORIES_FLG', 'CESS_FLG', 'THIRD_PARTY_FLG', 'AR4_FLG', 'REWARD_FLG', 'TOTAL_VAL_FC',
        'STR_FLG', 'END_USE', 'IGST_PAYMENT_STATUS', 'TAXABLE_VALUE', 'IGST_AMOUNT',
        'SWI_STO', 'SWI_DOO', 'SWI_EPT', 'SWI_UQC', 'SWI_QTY', 'SWI_GCESS_AMT', 'SWI_GCESS_CUR',
        'RODTEP_FLG', 'SOURCE_STATE', 'DBK_SRNO', 'DBK_QUANTITY'
    ]

    return output_df[final_columns]


# 💾 Save to Excel buffer (for download)
def save_to_excel_buffer(df, sheet_name="FormattedInvoice"):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)
    return output


def save_to_excel_97_2003_buffer(df, sheet_name="FormattedInvoice"):
    """
    Converts DataFrame to Excel 97-2003 (.xls) format and returns as a buffer
    """
    output = io.BytesIO()

    # Create workbook and worksheet
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet(sheet_name)

    # Write headers
    for col_idx, col_name in enumerate(df.columns):
        worksheet.write(0, col_idx, str(col_name))

    # Write data
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row):
            # Handle different types of data
            if pd.isna(value):
                worksheet.write(row_idx + 1, col_idx, '')
            else:
                worksheet.write(row_idx + 1, col_idx, value)

    # Save to buffer
    workbook.save(output)
    output.seek(0)
    return output
# Updated Streamlit code section for Bhajan

# Updated Streamlit code section for Bhajan
if selected_company == "Bhajan":
    uploaded_file = st.file_uploader("Upload Bhajan Invoice", type=["xlsx", "xls"])

    if uploaded_file:
        st.success("File Uploaded Successfully!")

        df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)

        # === Load Country Mapping ===
        try:
            country_df = pd.read_excel(COUNTRY_CODE_PATH)
            country_df['Country'] = country_df['Country'].astype(str).str.strip().str.upper()
            country_to_code = dict(zip(country_df['Country'], country_df['Code']))
            country_to_code['ANYTHING ELSE'] = 'NCPTI'
        except Exception as e:
            st.error(f"❌ Failed to load country code mapping: {e}")
            country_to_code = {}

        # === Detect Country in Raw Data ===
        detected_country = None
        for i in range(min(40, len(df_raw))):
            row = df_raw.iloc[i]
            for j in range(len(row) - 1):
                if isinstance(row[j], str) and 'country' in row[j].lower():
                    detected_country = str(row[j + 1]).strip().upper()
                    break
            if detected_country:
                break

        st.success(f"🌍 Detected Country: {detected_country if detected_country else 'NOT FOUND'}")

        # === Get SWI_EPT Code ===
        swi_ept_code = country_to_code.get(detected_country, country_to_code.get('ANYTHING ELSE', 'NCPTI'))
        swi_ept_code = st.text_input("📄 SWI_EPT Code (editable)", value=swi_ept_code)

        # --- Header Detection ---
        header_keywords = ["Sr. No.", "Product", "Description", "HSN", "Qty", "Rate", "Amount"]
        header_row_index = None

        for i, row in df_raw.iterrows():
            match_count = sum(
                any(k.lower() in str(cell).lower() for cell in row if pd.notna(cell)) for k in header_keywords)
            if match_count >= 3:
                header_row_index = i
                break

        # --- Process if Header is Found ---
        if header_row_index is not None:
            # st.info(f"Header row detected at index: {header_row_index}")
            df = pd.read_excel(uploaded_file, sheet_name=0, header=header_row_index)
            df.columns = [str(c).strip() for c in df.columns]

            # --- PRE-PROCESSING PIPELINE ---
            # 1. Remove completely blank rows
            df.dropna(how='all', inplace=True)
            df.reset_index(drop=True, inplace=True)

            trailing_keywords = ['less discount', 'discount', 'adjustment', 'deduction']

            rows_to_drop = []
            for idx, row in df.iterrows():
                row_str = ' '.join(row.dropna().astype(str)).lower()
                if any(k in row_str for k in trailing_keywords):
                    rows_to_drop.append(idx)

            if rows_to_drop:
                st.info(f"🧹 Removing {len(rows_to_drop)} trailing rows containing discount/adjustment summaries.")
                df.drop(rows_to_drop, inplace=True)
                df.reset_index(drop=True, inplace=True)

            # 2. Truncate the footer
            footer_keywords = [
                'round off', 'total invoice amount', 'in words', 'bank details',
                'bank name', 'bank a/c', 'certified that', 'authorised signatory',
                'common seal', 'total amount before tax', 'gst on reverse', 'add packing','Less Discount'
            ]
            footer_start_index = None
            for index, row in df.iterrows():
                row_as_string = ' '.join(row.dropna().astype(str)).lower()
                if any(keyword in row_as_string for keyword in footer_keywords):
                    footer_start_index = index
                    break

            if footer_start_index is not None:
                # st.info(f"Footer section detected starting at row {footer_start_index}. Truncating data...")
                df = df.iloc[:footer_start_index]
                df.reset_index(drop=True, inplace=True)

            # --- 3. Remove "Descriptive Header" Rows ---
            rows_to_drop = []
            key_value_columns = ['HSN', 'Qty.', 'Rate', 'in USD', 'Taxable value']

            for index, row in df.iterrows():
                is_sr_no_missing = pd.isna(row.get('Sr. No.'))
                existing_key_cols = [col for col in key_value_columns if col in df.columns]
                are_values_missing = row[existing_key_cols].isnull().all()

                if is_sr_no_missing and are_values_missing:
                    rows_to_drop.append(index)

            if rows_to_drop:
                st.info(f"Found and removed {len(rows_to_drop)} descriptive header rows that were not actual items.")
                df.drop(rows_to_drop, inplace=True)
                df.reset_index(drop=True, inplace=True)

            st.subheader("Detected Invoice Table (After All Cleaning)")
            st.dataframe(df)

            # --- Add option to simplify product descriptions ---
            simplify_descriptions = st.checkbox("Keep only main product name (remove text after first '/')", value=True)

            # --- CALL THE EXPORT FORMAT FUNCTION WITH SIMPLIFIED DESCRIPTIONS ---
            st.subheader("✨ Export Format(Consolidated Descriptions)")
            with st.spinner("Converting to export format..."):
                export_df2 = process_bhajan_to_export_format_alternative(df, simplify_descriptions)
            st.dataframe(export_df2)

            st.success("✅ Export format conversion complete!")

            chosen_df = export_df2

            # Add download button for the export format
            st.subheader("📦 Final Mapped Excel Format")
            dollar_price = st.number_input("💵 Enter Dollar Price (Exchange Rate)", min_value=0.0, value=83.0, step=0.1)
            if st.button("🔄 Generate Final Excel Table"):
                final_mapped_df = create_final_mapped_excel(
                    export_df2,
                    dollar_price,
                    uqc_mapping_path="/Users/piyanshu/PycharmProjects/pdftoexcel/SQC_1.xlsx",
                    swi_ept_code=swi_ept_code
                )
                st.session_state["final_df"] = final_mapped_df

                # Convert to Excel 97-2003 (.xls) format and store in session
                excel_buffer = save_to_excel_97_2003_buffer(final_mapped_df, sheet_name="FormattedInvoice")
                st.session_state["excel_output"] = excel_buffer

                st.success("✅ Final Excel 97-2003 Format ready!")
                st.dataframe(final_mapped_df)

            # === SAVE TO FOLDER ===
            if "excel_output" in st.session_state and st.button("📁 Add to Folder"):
                SAVE_DIRECTORY = "/Users/piyanshu/PycharmProjects/pdftoexcel/final_exports"
                os.makedirs(SAVE_DIRECTORY, exist_ok=True)

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_path = os.path.join(SAVE_DIRECTORY,
                                        f"formatted_invoice_bhajan_{timestamp}.xls")  # .xls extension for 97-2003 format
                try:
                    with open(file_path, "wb") as f:
                        f.write(st.session_state["excel_output"].getbuffer())
                    st.success(f"✅ File saved to: {file_path}")
                except Exception as e:
                    st.error(f"❌ Failed to save file: {e}")

            # === DOWNLOAD BUTTON ===
            if "excel_output" in st.session_state:
                st.download_button(
                    label="📥 Download Final Excel Table (97-2003 Format)",
                    data=st.session_state["excel_output"],
                    file_name="bhajana_invoice.xls",  # .xls extension for 97-2003 format
                    mime="application/vnd.ms-excel"  # MIME type for .xls
                )
        else:
            st.error("Could not detect a standard header row.")# bhajana ends




#pachranga start

import pandas as pd
import streamlit as st
import io
import os
from datetime import datetime
import xlwt  # For Excel 97-2003 format


def structure_pachranga_invoice(raw_df):
    """
    Extracts structured item data from Pachranga invoice DataFrame.
    Stops processing before declaration/footer like GR.WT., NET WT, Carton, Origin Declaration, etc.
    """
    target_columns = [
        'CARTON NO.', 'NO OF CTN', 'CBM', 'Item Name', 'HSN CODE', 'CARTON TO',
        'SIZE', 'QTY', 'TOTAL PCS', 'RATE (USD)', 'TOTAL USD'
    ]

    start_index = None
    for i, row in raw_df.iterrows():
        if any('CARTON NO' in str(cell).upper() for cell in row):
            start_index = i + 1
            break

    if start_index is None:
        st.error("❌ 'CARTON NO.' header not found in invoice.")
        return pd.DataFrame(columns=target_columns)

    data_rows = raw_df.iloc[start_index:].dropna(how='all')
    structured_data = []

    stop_keywords = [
        "GR. WT", "NET WT", "TOTAL PCS", "CARTON", "EXPORTER", "ORIGIN CRITERION", "THE PRODUCT COVERED", "DECLARATION",
        "US $", "INDIAN PREFERENTIAL ORIGIN"
    ]

    for _, row in data_rows.iterrows():
        values = row.dropna().astype(str).tolist()

        # 🚫 Stop if footer/declaration content is detected
        if any(any(keyword in v.upper() for keyword in stop_keywords) for v in values):
            break

        # Skip non-item rows
        if any("SUPPLY MEANT" in str(v).upper() or "DISCOUNT" in str(v).upper() for v in values):
            continue

        if len(values) >= 12:
            try:
                carton_raw = str(values[0])
                carton_no = carton_raw.split()[0] if 'TO' in carton_raw.upper() else carton_raw

                structured_data.append({
                    'CARTON NO.': carton_no,
                    'CARTON TO': values[2],
                    'NO OF CTN': values[3],
                    'CBM': values[4],
                    'Item Name': values[5],
                    'HSN CODE': values[6],
                    'SIZE': values[7],
                    'QTY': values[8],
                    'TOTAL PCS': values[9],
                    'RATE (USD)': values[10],
                    'TOTAL USD': values[11]
                })
            except Exception as e:
                st.warning(f"⚠️ Skipping row due to error: {e}")
                continue

    return pd.DataFrame(structured_data, columns=target_columns)


def clean_total_usd_value(value):
    """
    Cleans TOTAL USD values by removing spaces, commas, and converting to float
    """
    if pd.isna(value):
        return 0.0

    value_str = str(value).strip()
    # Remove spaces, commas, and other non-numeric characters except decimal point
    cleaned_value = ''.join(char for char in value_str if char.isdigit() or char == '.')

    try:
        return float(cleaned_value) if cleaned_value else 0.0
    except ValueError:
        return 0.0


def to_excel_97_2003_buffer(df, sheet_name="StructuredData"):
    """
    Converts DataFrame to Excel 97-2003 (.xls) format and returns as a buffer
    """
    output = io.BytesIO()

    # Create workbook and worksheet
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet(sheet_name)

    # Write headers
    for col_idx, col_name in enumerate(df.columns):
        worksheet.write(0, col_idx, col_name)

    # Write data
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row):
            # Handle different types of data
            if pd.isna(value):
                worksheet.write(row_idx + 1, col_idx, '')
            elif isinstance(value, (int, float)):
                # Write numeric values as numbers
                worksheet.write(row_idx + 1, col_idx, value)
            else:
                # Write other values as strings
                worksheet.write(row_idx + 1, col_idx, str(value))

    # Save to buffer
    workbook.save(output)
    output.seek(0)
    return output


def create_final_mapped_excel_pachranga(structured_df, swi_ept_code="NCPTI"):
    final_rows = []

    for i, row in structured_df.iterrows():
        # Keep original RATE_VALUE exactly as it is in structured invoice
        rate_usd_original = row.get("RATE (USD)", "0")

        # Clean TOTAL USD value for TOTAL_VAL_FC
        total_usd_raw = str(row.get("TOTAL USD", "0")).strip()
        total_usd = clean_total_usd_value(total_usd_raw)

        final_rows.append({
            "INVOICE_SR_NO": 1,
            "ITEM_SR_NO": i + 1,
            "SCHEME_CODE": 19,
            "RITC": row.get("HSN CODE", ""),
            "GOODS_DESC1": row.get("Item Name", ""),
            "GOODS_DESC2": "PACK:",
            "GOODS_DESC3": "",
            "QTY_NOS": row.get("TOTAL PCS", 0),
            "UNIT_QTY": "PCS",
            "RATE_VALUE": rate_usd_original,  # Exact original value preserved
            "NO_OF_UNIT": 1,
            "UNIT_OF_RATE": "PCS",
            "PMV_AMT": "",
            "TOTAL_PMV": "",
            "ACCESSORIES_FLG": "N",
            "CESS_FLG": "N",
            "THIRD_PARTY_FLG": "N",
            "AR4_FLG": "N",
            "REWARD_FLG": "Y",
            "TOTAL_VAL_FC": total_usd,  # Properly cleaned numeric value
            "STR_FLG": "N",
            "END_USE": "GNX100",
            "IGST_PAYMENT_STATUS": "LUT",
            "TAXABLE_VALUE": 0,
            "IGST_AMOUNT": 0,
            "SWI_STO": "06",
            "SWI_DOO": "71",
            "SWI_EPT": swi_ept_code,
            "SWI_UQC": "KGS",
            "SWI_QTY": "",
            "SWI_GCESS_AMT": 0,
            "SWI_GCESS_CUR": "INR",
            "RODTEP_FLG": "Y",
            "SOURCE_STATE": "06",
            "DBK_SRNO": "",
            "DBK_QUANTITY": ""
        })

    return pd.DataFrame(final_rows)


# === Streamlit UI for Pachranga Foods ===
if selected_company == "PACHRANGA FOODS":
    uploaded_file = st.file_uploader("📤 Upload PACHRANGA FOODS Invoice", type=["xlsx", "xls"])

    if uploaded_file:
        st.success("📄 File Uploaded Successfully!")

        df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
        df_raw = df_raw[
            ~df_raw.apply(lambda row: row.astype(str).str.contains("Amount Chargable", case=False, na=False)).any(
                axis=1)]

        st.subheader("📄 Raw Invoice Preview")
        st.dataframe(df_raw)

        structured_df = structure_pachranga_invoice(df_raw)

        if not structured_df.empty:
            st.subheader("🧾 Structured Invoice Data")
            st.dataframe(structured_df)

            try:
                # Load while specifying header row
                country_df = pd.read_excel(COUNTRY_CODE_PATH, header=0, usecols=["Country", "Code"])
                country_df['Country'] = country_df['Country'].astype(str).str.strip().str.upper()
                country_to_code = dict(zip(country_df['Country'], country_df['Code']))
                country_to_code['ANYTHING ELSE'] = 'NCPTI'
            except Exception as e:
                st.error(f"❌ Failed to load country code mapping: {e}")
                country_to_code = {}

            detected_country = None
            for i in range(min(40, len(df_raw))):
                row = df_raw.iloc[i]
                for j in range(len(row) - 1):
                    if isinstance(row[j], str) and 'country' in row[j].lower():
                        detected_country = str(row[j + 1]).strip().upper()
                        break
                if detected_country:
                    break

            st.success(f"🌍 Detected Country: {detected_country if detected_country else 'NOT FOUND'}")

            swi_ept_code = country_to_code.get(detected_country, country_to_code.get('ANYTHING ELSE', 'NCPTI'))
            swi_ept_code = st.text_input("📄 SWI_EPT Code (editable)", value=swi_ept_code)

            if st.button("🔄 Generate Final Excel Table"):
                final_mapped_df = create_final_mapped_excel_pachranga(structured_df, swi_ept_code)
                st.session_state["final_df"] = final_mapped_df

                # Create Excel 97-2003 (.xls) file using xlwt
                excel_buffer = to_excel_97_2003_buffer(final_mapped_df, sheet_name="FormattedInvoice")
                st.session_state["excel_output"] = excel_buffer

                st.success("✅ Final Excel (97-2003 format) ready!")
                st.dataframe(final_mapped_df)

            if "excel_output" in st.session_state and st.button("📁 Add to Folder"):
                SAVE_DIRECTORY = "/Users/piyanshu/PycharmProjects/pdftoexcel/final_exports"
                os.makedirs(SAVE_DIRECTORY, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_path = os.path.join(SAVE_DIRECTORY,
                                         f"formatted_invoice_pachranga_{timestamp}.xls")  # .xls extension for 97-2003 format
                try:
                    with open(file_path, "wb") as f:
                        f.write(st.session_state["excel_output"].getbuffer())
                    st.success(f"✅ File saved to: {file_path}")
                except Exception as e:
                    st.error(f"❌ Failed to save file: {e}")

            if "excel_output" in st.session_state:
                st.download_button(
                    label="📥 Download Final Excel Table (97-2003 Format)",
                    data=st.session_state["excel_output"],
                    file_name="final_pachranga_invoice.xls",  # .xls extension for 97-2003 format
                    mime="application/vnd.ms-excel"  # MIME type for .xls
                )
        else:
            st.warning("⚠️ No valid structured data extracted.")

#-------Pachranga finish--------



import streamlit as st
import xlwt
import pandas as pd

# imperial starts
from decimal import Decimal, getcontext, ROUND_HALF_UP

getcontext().prec = 10  # or more if needed

def precise_mul(qty, rate):
    try:
        qty_dec = Decimal(str(qty))
        rate_dec = Decimal(str(rate))
        result = qty_dec * rate_dec
        return result.quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)  # round to 4 decimal places
    except:
        return Decimal('0.0000')

def create_final_mapped_excel_imperial(export_df2, dollar_price, uqc_mapping_path=None, swi_ept_code="NCPTI"):
    """
    Generates the final mapped Excel DataFrame for Imperial invoices,
    ensuring numeric calculations and field mapping.
    """
    ritc_to_uqc = {}

    if uqc_mapping_path:
        try:
            uqc_map_df = pd.read_excel(uqc_mapping_path)
            uqc_map_df['RITC'] = uqc_map_df['RITC'].astype(str).str.strip()
            ritc_to_uqc = dict(zip(uqc_map_df['RITC'], uqc_map_df['SWI_UQC']))
        except Exception as e:
            st.error(f"⚠️ Failed to load UQC mapping: {e}")
            ritc_to_uqc = {}

    output_df = pd.DataFrame()
    output_df['INVOICE_SR_NO'] = [1] * len(export_df2)
    output_df['ITEM_SR_NO'] = export_df2['Sr. No.']
    output_df['SCHEME_CODE'] = export_df2['HSN Code'].astype(str).apply(
        lambda x: "60" if x.startswith(("60", "61", "62", "63")) else "19")
    output_df['RITC'] = export_df2['HSN Code']
    output_df['GOODS_DESC1'] = export_df2['Product']
    output_df['GOODS_DESC2'] = "SIZE:"
    output_df['GOODS_DESC3'] = "As Per Invoice"
    output_df['QTY_NOS'] = export_df2['Qty (in pcs)']
    output_df['UNIT_QTY'] = export_df2['Unit_qty']
    output_df['RATE_VALUE'] = export_df2['Exchange Rate (in usd/  Pcs)']
    output_df['NO_OF_UNIT'] = 1
    output_df['UNIT_OF_RATE'] = output_df['UNIT_QTY']
    output_df['PMV_AMT'] = ""
    output_df['TOTAL_PMV'] = ""
    output_df['ACCESSORIES_FLG'] = "N"
    output_df['CESS_FLG'] = "N"
    output_df['THIRD_PARTY_FLG'] = "N"
    output_df['AR4_FLG'] = "N"
    output_df['REWARD_FLG'] = "Y"
    output_df['TOTAL_VAL_FC'] = export_df2['Taxable value(in USD)']
    output_df['STR_FLG'] = "N"
    output_df['END_USE'] = "GNX100"
    output_df['IGST_PAYMENT_STATUS'] = "P"
    output_df['TAXABLE_VALUE'] = pd.to_numeric(output_df['TOTAL_VAL_FC'], errors='coerce')
    # --- Fix numeric conversion and compute correct amounts ---
    output_df['TOTAL_VAL_FC'] = export_df2.apply(
        lambda row: precise_mul(row['Qty (in pcs)'], row['Exchange Rate (in usd/  Pcs)']), axis=1
    )
    igst_rate = pd.to_numeric(export_df2['IGST Rate (%)'], errors='coerce')


    output_df['IGST_AMOUNT'] = pd.to_numeric(export_df2['IGST Amount (in Rs)'], errors='coerce').round(2)

    output_df['SWI_STO'] = "06"
    output_df['SWI_DOO'] = "71"
    output_df['SWI_EPT'] = swi_ept_code
    output_df['SWI_UQC'] = output_df['RITC'].astype(str).str.strip().map(ritc_to_uqc).fillna("NOS")
    output_df['SWI_QTY'] = ""
    output_df['SWI_GCESS_AMT'] = 0
    output_df['SWI_GCESS_CUR'] = "INR"
    output_df['RODTEP_FLG'] = output_df['SCHEME_CODE'].apply(lambda x: "N" if x == "60" else "Y")
    output_df['SOURCE_STATE'] = "06"
    output_df['DBK_SRNO'] = ""
    output_df['DBK_QUANTITY'] = ""

    final_columns = [
        'INVOICE_SR_NO', 'ITEM_SR_NO', 'SCHEME_CODE', 'RITC', 'GOODS_DESC1', 'GOODS_DESC2', 'GOODS_DESC3',
        'QTY_NOS', 'UNIT_QTY', 'RATE_VALUE', 'NO_OF_UNIT', 'UNIT_OF_RATE', 'PMV_AMT', 'TOTAL_PMV',
        'ACCESSORIES_FLG', 'CESS_FLG', 'THIRD_PARTY_FLG', 'AR4_FLG', 'REWARD_FLG', 'TOTAL_VAL_FC',
        'STR_FLG', 'END_USE', 'IGST_PAYMENT_STATUS', 'TAXABLE_VALUE', 'IGST_AMOUNT',
        'SWI_STO', 'SWI_DOO', 'SWI_EPT', 'SWI_UQC', 'SWI_QTY', 'SWI_GCESS_AMT', 'SWI_GCESS_CUR',
        'RODTEP_FLG', 'SOURCE_STATE', 'DBK_SRNO', 'DBK_QUANTITY'
    ]

    return output_df[final_columns]


def clean_numeric_column(col):
    """Clean numeric columns by removing currency symbols and commas"""
    if col.dtype == 'object':
        # Remove currency symbols, commas, and extract numeric part
        cleaned = col.astype(str).str.replace('₹', '').str.replace(',', '').str.strip()
        return pd.to_numeric(cleaned, errors='coerce')
    return pd.to_numeric(col, errors='coerce')


def extract_main_table_from_imperial_invoice(raw_df):
    """
    Extracts the main item table from Imperial invoice with the new format.
    Handles the specific structure where data is arranged in rows.
    Fixed to properly map IGST Amount and Taxable Value columns.
    """
    print("Debug: Starting extraction...")

    # Convert all data to string and handle NaN values
    df_str = raw_df.fillna('').astype(str)

    # Print first few rows for debugging
    print("Debug: First 10 rows of raw data:")
    for i in range(min(10, len(df_str))):
        print(f"Row {i}: {df_str.iloc[i].tolist()}")

    # Find where the data starts (look for product descriptions)
    data_start_idx = None
    product_keywords = ['COTTON', 'RUGS', 'POUF', 'CUSHION']

    for idx, row in df_str.iterrows():
        row_text = ' '.join(row.values).upper()
        if any(keyword in row_text for keyword in product_keywords):
            data_start_idx = idx
            print(f"Debug: Found data start at row {idx}: {row_text}")
            break

    if data_start_idx is None:
        st.error("Could not find data start in the invoice")
        return pd.DataFrame()

    # Extract data starting from identified row
    data_rows = []
    current_description = ""

    # Define expected columns
    expected_columns = [
        'S NO.', 'DESCRIPTION OF GOODS', 'HSN', 'QTY', 'UNIT',
        'RATE PER PC', 'TOTAL', 'FREIGHT', 'TAXABLE VALUE',
        'IGST RATE', 'IGST AMOUNT', 'TOTAL VALUE'
    ]

    # Process rows from data start
    for idx in range(data_start_idx, len(df_str)):
        row = df_str.iloc[idx]
        row_values = [cell.strip() for cell in row.values if cell.strip() and cell.strip() != 'nan']

        if not row_values:
            continue

        print(f"Debug: Processing row {idx}: {row_values}")

        # Check if this is a footer row
        row_text = ' '.join(row_values).lower()
        footer_keywords = ['total igst', 'total invoice', 'grand total', 'signature', 'authorised signatory', 'cartons']
        if any(keyword in row_text for keyword in footer_keywords):
            print(f"Debug: Found footer at row {idx}, stopping extraction")
            break

        # Check if this row contains a product description
        is_description_row = False
        for keyword in product_keywords:
            if keyword in row_text.upper() and len(row_values) <= 3:
                current_description = ' '.join(row_values)
                is_description_row = True
                print(f"Debug: Found description: {current_description}")
                break

        if is_description_row:
            continue

        # Check if this row contains data (has HSN code - 8 digits)
        hsn_value = ""
        qty_value = ""
        unit_value = ""
        rate_value = ""
        total_value = ""
        freight_value = ""
        taxable_value = ""
        igst_rate = ""
        igst_amount = ""
        total_final = ""

        # Look for HSN code (8 digits)
        for val in row_values:
            if val.isdigit() and len(val) == 8:
                hsn_value = val
                print(f"Debug: Found HSN: {hsn_value}")
                break

        # If we found HSN, extract other values from this row
        if hsn_value and current_description:
            # Parse the row based on expected positions
            try:
                hsn_idx = row_values.index(hsn_value)
                remaining_values = row_values[hsn_idx + 1:]
                print(f"Debug: Values after HSN: {remaining_values}")

                # Extract QTY and UNIT first
                if len(remaining_values) >= 2:
                    qty_value = remaining_values[0] if remaining_values[0].replace(',', '').isdigit() else ""
                    unit_value = remaining_values[1] if remaining_values[1].upper() in ['PCS', 'PC', 'PIECE',
                                                                                        'PIECES'] else ""

                # Separate different types of values
                monetary_values = []
                decimal_values = []

                # Process remaining values - start from index 2 (after QTY and UNIT)
                for val in remaining_values[2:]:
                    # Check for IGST rate (decimal like 0.05, 0.18)
                    if val.replace('.', '').replace('0', '').isdigit() and '0.' in val and len(val) <= 4:
                        decimal_values.append(val)
                    # Check for monetary values (containing ₹ or formatted numbers)
                    elif '₹' in val or (',' in val and any(c.isdigit() for c in val)):
                        monetary_values.append(val)
                    # Check for simple numeric values that could be monetary values
                    elif val.replace('.', '').replace(',', '').isdigit():
                        monetary_values.append(val)

                print(f"Debug: Monetary values: {monetary_values}")
                print(f"Debug: Decimal values (IGST rates): {decimal_values}")

                # Assign IGST rate first
                if decimal_values:
                    igst_rate = decimal_values[0]

                # Now assign monetary values in the correct order
                # IMPORTANT FIX: Check for the exact number of values to correctly identify RATE PER PC
                if len(monetary_values) >= 1:
                    # The first monetary value is RATE PER PC
                    rate_value = monetary_values[0]

                if len(monetary_values) >= 2:
                    # The second monetary value is TOTAL
                    total_value = monetary_values[1]

                if len(monetary_values) >= 3:
                    # If we have many values, use the following logic:
                    if len(monetary_values) >= 5:
                        # We have all values: RATE, TOTAL, FREIGHT, TAXABLE, IGST, TOTAL_FINAL
                        freight_value = monetary_values[2]
                        taxable_value = monetary_values[3]
                        igst_amount = monetary_values[4]
                        if len(monetary_values) >= 6:
                            total_final = monetary_values[5]
                    elif len(monetary_values) == 4:
                        # We might be missing FREIGHT or IGST
                        # Calculate expected values to determine which one is missing
                        try:
                            rate_float = float(rate_value.replace('₹', '').replace(',', ''))
                            qty_float = float(qty_value.replace(',', ''))
                            calculated_total = rate_float * qty_float

                            # Check if any value matches the calculated total
                            total_match = False
                            for i, val in enumerate(monetary_values):
                                val_float = float(val.replace('₹', '').replace(',', ''))
                                if abs(val_float - calculated_total) < (calculated_total * 0.1):  # 10% tolerance
                                    total_match = True
                                    total_value = val
                                    # The value after total should be taxable
                                    if i + 1 < len(monetary_values):
                                        taxable_value = monetary_values[i + 1]
                                    # The value after taxable should be IGST amount
                                    if i + 2 < len(monetary_values):
                                        igst_amount = monetary_values[i + 2]
                                    break

                            # If no match found, follow the standard pattern
                            if not total_match:
                                freight_value = ""
                                taxable_value = monetary_values[2]
                                igst_amount = monetary_values[3]

                        except (ValueError, IndexError):
                            # Fallback to simple mapping if calculation fails
                            freight_value = ""
                            taxable_value = monetary_values[2]
                            igst_amount = monetary_values[3]
                    else:
                        # With only 3 values, we're likely missing both FREIGHT and IGST
                        freight_value = ""
                        taxable_value = monetary_values[2]
                        igst_amount = ""

                # Additional validation for IGST calculation
                if taxable_value and igst_rate:
                    try:
                        clean_taxable = float(taxable_value.replace('₹', '').replace(',', ''))
                        clean_igst_rate = float(igst_rate)
                        expected_igst = clean_taxable * clean_igst_rate

                        # If we have an IGST amount, validate it
                        if igst_amount:
                            clean_igst_amount = float(igst_amount.replace('₹', '').replace(',', ''))
                            # If there's a significant mismatch
                            if abs(expected_igst - clean_igst_amount) > (clean_taxable * 0.01):
                                print(
                                    f"Warning: IGST calculation mismatch. Expected: {expected_igst}, Got: {clean_igst_amount}")
                                # Try to find a value that matches the expected IGST
                                for val in monetary_values:
                                    try:
                                        test_val = float(val.replace('₹', '').replace(',', ''))
                                        if abs(expected_igst - test_val) <= (clean_taxable * 0.01):
                                            igst_amount = val
                                            break
                                    except:
                                        continue
                    except:
                        pass

                # Create data row
                data_row = [
                    '',  # S NO. - will be filled later
                    current_description,  # DESCRIPTION OF GOODS
                    hsn_value,  # HSN
                    qty_value,  # QTY
                    unit_value,  # UNIT
                    rate_value,  # RATE PER PC
                    total_value,  # TOTAL
                    freight_value,  # FREIGHT
                    taxable_value,  # TAXABLE VALUE
                    igst_rate,  # IGST RATE
                    igst_amount,  # IGST AMOUNT
                    total_final  # TOTAL VALUE
                ]

                print(f"Debug: Created data row: {data_row}")
                data_rows.append(data_row)

            except Exception as e:
                print(f"Debug: Error processing row {idx}: {e}")
                continue

    # Create DataFrame
    if data_rows:
        df = pd.DataFrame(data_rows, columns=expected_columns)
        # Add serial numbers
        df['S NO.'] = range(1, len(df) + 1)
        print(f"Debug: Created DataFrame with {len(df)} rows")

        # Final validation and cleanup
        df = validate_and_fix_column_mapping(df)

        return df
    else:
        print("Debug: No data rows found")
        return pd.DataFrame()

def validate_and_fix_column_mapping(df):
    """
    Validates and fixes any remaining column mapping issues.
    """
    if df.empty:
        return df

    print("Debug: Validating column mapping...")

    for idx, row in df.iterrows():
        try:
            # Check if IGST calculation is correct
            taxable_val = str(row['TAXABLE VALUE']).replace('₹', '').replace(',', '').strip()
            igst_rate = str(row['IGST RATE']).replace('₹', '').replace(',', '').strip()
            igst_amount = str(row['IGST AMOUNT']).replace('₹', '').replace(',', '').strip()

            if taxable_val and igst_rate and igst_amount:
                try:
                    taxable_num = float(taxable_val)
                    rate_num = float(igst_rate)
                    amount_num = float(igst_amount)

                    expected_amount = taxable_num * rate_num

                    # If there's a significant mismatch, flag it
                    if abs(expected_amount - amount_num) > (taxable_num * 0.02):  # 2% tolerance
                        print(f"Row {idx}: IGST mismatch detected. Expected: {expected_amount}, Got: {amount_num}")

                        # Try to find values that look like they're in wrong columns
                        # Check if IGST AMOUNT value is actually in TAXABLE VALUE
                        if amount_num > taxable_num:  # IGST amount shouldn't be larger than taxable value
                            print(f"Row {idx}: Swapping TAXABLE VALUE and IGST AMOUNT")
                            df.at[idx, 'TAXABLE VALUE'] = igst_amount
                            df.at[idx, 'IGST AMOUNT'] = taxable_val

                except ValueError:
                    print(f"Row {idx}: Could not convert values for validation")
                    continue

        except Exception as e:
            print(f"Row {idx}: Validation error: {e}")
            continue

    return df

def clean_description_and_drop_sparse_rows(df, max_empty_allowed=2):
    """
    Clean the extracted data and handle sparse rows.
    Also removes rows where QTY is 0.
    """
    if df.empty:
        return df

    df = df.copy()

    # Clean empty strings and normalize
    df = df.replace('', pd.NA)
    df = df.dropna(how='all')

    # Drop rows with too many empty cells
    df = df[df.isnull().sum(axis=1) <= max_empty_allowed]

    # Convert QTY to numeric for filtering
    if 'QTY' in df.columns:
        df['QTY'] = pd.to_numeric(df['QTY'], errors='coerce')
        df = df[df['QTY'] != 0]  # ❌ Remove rows with QTY == 0

    if 'DESCRIPTION OF GOODS' in df.columns:
        df['DESCRIPTION OF GOODS'] = df['DESCRIPTION OF GOODS'].astype(str).str.rstrip('0').str.strip()

    # Reset index and update serial numbers
    df.reset_index(drop=True, inplace=True)
    df['S NO.'] = range(1, len(df) + 1)

    return df


def process_imperial_to_export_format(structured_df, dollar_price=83.0):
    """
    Process cleaned Imperial structured data to export format.
    """
    if structured_df.empty:
        return pd.DataFrame()

    df = structured_df.copy()

    # Map columns to export format
    export_df = pd.DataFrame()
    export_df["Sr. No."] = df["S NO."]
    export_df["Product"] = df["DESCRIPTION OF GOODS"]
    export_df["HSN Code"] = df["HSN"]
    export_df["Qty (in pcs)"] = pd.to_numeric(df["QTY"], errors='coerce')
    export_df["Unit_qty"] = df["UNIT"]

    # Make sure RATE PER PC is properly extracted
    export_df["Exchange Rate (in usd/  Pcs)"] = clean_numeric_column(df["RATE PER PC"])

    # Calculate 'in USD' from rate and quantity as a double-check
    export_df["in USD"] = export_df["Qty (in pcs)"] * export_df["Exchange Rate (in usd/  Pcs)"]

    # Compare with the extracted TOTAL and use the extracted if they're close
    extracted_total = clean_numeric_column(df["TOTAL"])
    for i, row in export_df.iterrows():
        calculated = row["in USD"]
        extracted = extracted_total.iloc[i] if i < len(extracted_total) else 0

        # If the extracted total is within 10% of calculated, use it
        if not pd.isna(extracted) and not pd.isna(calculated) and abs(extracted - calculated) < (calculated * 0.1):
            export_df.at[i, "in USD"] = extracted

    export_df["Taxable value(in USD)"] = clean_numeric_column(df["TAXABLE VALUE"])
    export_df["IGST Rate (%)"] = clean_numeric_column(df["IGST RATE"])
    export_df["IGST Amount (in Rs)"] = clean_numeric_column(df["IGST AMOUNT"])
    export_df["Total (in Rs)"] = clean_numeric_column(df["TOTAL VALUE"])

    # Clean the dataframe
    export_df = export_df.dropna(subset=["HSN Code", "Product"])

    return export_df

# Country mapping functions (unchanged)
COUNTRY_CODE_PATH = "/Users/piyanshu/PycharmProjects/pdftoexcel/Country Code.xlsx"


def load_country_mapping():
    try:
        country_df = pd.read_excel(COUNTRY_CODE_PATH)
        country_df['Country'] = country_df['Country'].astype(str).str.strip().str.upper()
        country_df['Code'] = country_df['Code'].astype(str).str.strip()
        mapping = dict(zip(country_df['Country'], country_df['Code']))
        mapping['ANYTHING ELSE'] = 'NCPTI'
        return mapping
    except Exception as e:
        st.error(f"❌ Failed to load country code mapping: {e}")
        return {'ANYTHING ELSE': 'NCPTI'}


def extract_country_from_invoice_text(raw_df, country_list):
    all_text = " ".join(raw_df.fillna('').astype(str).values.flatten()).upper()
    for country in country_list:
        if country in all_text:
            return country
    return "ANYTHING ELSE"


# Streamlit section (updated)
# Streamlit section (updated for Excel 97-2003 format)
if selected_company == "Imperial":
    uploaded_file = st.file_uploader("📤 Upload IMPERIAL Invoice", type=["xlsx", "xls", "csv"])

    if uploaded_file:
        st.success("✅ File uploaded successfully!")

        # Read raw file
        if uploaded_file.name.endswith(".csv"):
            df_raw = pd.read_csv(uploaded_file, header=None)
        else:
            df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)

        # Load Country Mapping
        country_to_code = load_country_mapping()

        # Extract Country
        detected_country = extract_country_from_invoice_text(df_raw, list(country_to_code.keys()))
        country_code = country_to_code.get(detected_country, "NCPTI")

        st.markdown(f"### 🌍 Destination Country: `{detected_country}`")
        st.markdown(f"### 🏷️ Country Code: `{country_code}`")
        swi_ept_code = st.text_input("📄 SWI_EPT Code (editable)", value=country_code)

        st.subheader("📄 Raw Invoice Preview")
        st.dataframe(df_raw)

        # Extract and clean main table
        structured_df = extract_main_table_from_imperial_invoice(df_raw)

        if not structured_df.empty:
            structured_df_cleaned = clean_description_and_drop_sparse_rows(structured_df, max_empty_allowed=2)

            st.subheader("🧾 Cleaned Structured Table")
            st.dataframe(structured_df_cleaned)
            dollar_price = st.number_input("💵 Enter Dollar Price (Exchange Rate)", min_value=0.0, value=83.0,
                                           step=0.1)

            # Convert to Export Format
            export_df2 = process_imperial_to_export_format(structured_df_cleaned, dollar_price)


            if not export_df2.empty:
                st.subheader("📦 Intermediate Export Format")
                st.dataframe(export_df2)

                # Input Dollar Price

                if st.button("🔄 Generate Final Excel Table"):
                    final_mapped_df = create_final_mapped_excel_imperial(
                        export_df2,
                        dollar_price=dollar_price,
                        uqc_mapping_path="/Users/piyanshu/PycharmProjects/pdftoexcel/SQC_1.xlsx",
                        swi_ept_code=swi_ept_code
                    )
                    st.session_state["final_df"] = final_mapped_df

                    # Convert to Excel 97-2003 format using xlwt
                    import io

                    output = io.BytesIO()
                    workbook = xlwt.Workbook()
                    worksheet = workbook.add_sheet("FormattedInvoice")

                    # Write headers
                    for col_idx, col_name in enumerate(final_mapped_df.columns):
                        worksheet.write(0, col_idx, str(col_name))

                    # Write data
                    for row_idx, row in final_mapped_df.iterrows():
                        for col_idx, value in enumerate(row):
                            if pd.isna(value):
                                worksheet.write(row_idx + 1, col_idx, '')
                            else:
                                worksheet.write(row_idx + 1, col_idx, value)

                    # Save to buffer
                    workbook.save(output)
                    output.seek(0)
                    st.session_state["excel_output"] = output

                    st.success("✅ Final Excel 97-2003 Format ready!")
                    st.dataframe(final_mapped_df)

                # Save to Folder
                if "excel_output" in st.session_state and st.button("📁 Add to Folder"):
                    import os
                    from datetime import datetime

                    SAVE_DIRECTORY = "/Users/piyanshu/PycharmProjects/pdftoexcel/final_exports"
                    os.makedirs(SAVE_DIRECTORY, exist_ok=True)

                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    file_path = os.path.join(SAVE_DIRECTORY,
                                             f"formatted_invoice_imperial_{timestamp}.xls")  # Changed to .xls
                    try:
                        with open(file_path, "wb") as f:
                            f.write(st.session_state["excel_output"].getbuffer())
                        st.success(f"✅ File saved to: {file_path}")
                    except Exception as e:
                        st.error(f"❌ Failed to save file: {e}")

                # Final Download Button
                if "excel_output" in st.session_state:
                    st.download_button(
                        label="📥 Download Final Excel Table (97-2003 Format)",
                        data=st.session_state["excel_output"],
                        file_name="final_imperial_invoice.xls",  # Changed to .xls
                        mime="application/vnd.ms-excel"  # Changed MIME type for .xls
                    )
            else:
                st.warning("⚠️ No valid data found in the structured table.")
        else:
            st.warning("⚠️ Could not detect the structured table from the invoice.")








###Bhajana SQm



#BHAJANA_SQM


def process_bhajan_to_export_format_alternative1(df, remove_after_slash=False):
    """
    Enhanced version that processes Bhajan invoices and maps additional fields like HSN, Qty, etc.
    Produces consolidated descriptions and all important numeric fields.

    Parameters:
    df - The input DataFrame
    remove_after_slash - If True, keeps only the part before the first slash in product descriptions
    """
    import pandas as pd

    df_copy = df.copy()

    # Clean and group
    df_copy['Sr. No.'] = pd.to_numeric(df_copy['Sr. No.'], errors='coerce')
    df_copy['Group_ID'] = df_copy['Sr. No.'].ffill()
    df_copy.dropna(subset=['Group_ID'], inplace=True)

    final_records = []
    output_sr_no = 1

    for _, group in df_copy.groupby('Group_ID'):
        parent_row = group.iloc[0]
        parent_description = str(parent_row.get('Product', '')).strip()

        # Gather sub-descriptions and rows with rates
        rate_rows = []
        sub_descriptions = []

        for i in range(1, len(group)):
            sub_row = group.iloc[i]

            desc = str(sub_row.get('Product', '')).strip()
            if desc:
                sub_descriptions.append(desc)

            if pd.notna(sub_row.get('Rate')) and pd.notna(sub_row.get('in USD')):
                rate_rows.append(sub_row)

        all_descriptions = [parent_description] + sub_descriptions
        consolidated_description = " / ".join(filter(None, all_descriptions))

        # Clean up product descriptions - keep only part before first slash if option is enabled
        if remove_after_slash and '/' in consolidated_description:
            # Split at the first forward slash and keep only the first part
            parts = consolidated_description.split('/', 1)
            consolidated_description = parts[0].strip()

        for rate_row in rate_rows:
            record = {
                'Sr. No.': output_sr_no,
                'Product': consolidated_description,
                'Unit_qty': rate_row.get('Unnamed: 5') or rate_row.get('In Pcs'),
                'in USD': rate_row.get('in USD'),
                'Taxable value(in USD)': rate_row.get('in USD'),  # Redundant but kept for clarity

                # Additional fields below
                'HSN Code': rate_row.get('HSN') or rate_row.get('HSN Code'),
                'Qty (in pcs)': rate_row.get('Unnamed: 4') or rate_row.get('Qty.'),
                'Qty (Sq./Mtrs)': rate_row.get('Sq.') or rate_row.get('Mtrs'),  # in usd/  Pcs
                'Exchange Rate (in usd/  Pcs)': rate_row.get('Rate') or rate_row.get('in usd/  Pcs'),
                'Taxable Value (in Rs)': rate_row.get('Taxable value.1'),
                'IGST Rate (%)': rate_row.get('IGST'),
                'IGST Amount (in Rs)': rate_row.get('Amount in Rs.'),
                'Total (in Rs)': rate_row.get('Total in Rs.')
            }

            final_records.append(record)
            output_sr_no += 1

    if not final_records:
        return pd.DataFrame()

    final_df = pd.DataFrame(final_records)

    # Final column order
    target_columns = [
        'Sr. No.', 'Product', 'Unit_qty', 'in USD', 'Taxable value(in USD)',
        'HSN Code', 'Qty (in pcs)', 'Qty (Sq./Mtrs)', 'Exchange Rate (in usd/  Pcs)',
        'Taxable Value (in Rs)', 'IGST Rate (%)', 'IGST Amount (in Rs)', 'Total (in Rs)'
    ]

    final_df = final_df[target_columns]

    return final_df.reset_index(drop=True)


def create_final_mapped_excel1(export_df2, dollar_price, uqc_mapping_path=None, swi_ept_code="NCPTI"):
    # Optional UQC Mapping
    ritc_to_uqc = {}
    if uqc_mapping_path:
        try:
            uqc_map_df = pd.read_excel(uqc_mapping_path)
            uqc_map_df['RITC'] = uqc_map_df['RITC'].astype(str).str.strip()
            ritc_to_uqc = dict(zip(uqc_map_df['RITC'], uqc_map_df['SWI_UQC']))
        except Exception as e:
            st.error(f"⚠️ Failed to load UQC mapping: {e}")
            ritc_to_uqc = {}

    output_df = pd.DataFrame()
    output_df['INVOICE_SR_NO'] = [1] * len(export_df2)
    output_df['ITEM_SR_NO'] = export_df2['Sr. No.']
    output_df['SCHEME_CODE'] = export_df2['HSN Code'].astype(str).apply(
        lambda x: "60" if x.startswith(("60", "61", "62", "63")) else "19")
    output_df['RITC'] = export_df2['HSN Code']
    output_df['GOODS_DESC1'] = export_df2['Product']
    output_df['GOODS_DESC2'] = "SIZE:"
    output_df['GOODS_DESC3'] = "As Per Invoice"

    # Keep exact original values - no conversion
    output_df['QTY_NOS'] = export_df2['Qty (Sq./Mtrs)']  # Exact original value
    output_df['UNIT_QTY'] = 'SQM'
    output_df['RATE_VALUE'] = export_df2['Exchange Rate (in usd/  Pcs)']  # Exact original value

    output_df['NO_OF_UNIT'] = 1
    output_df['UNIT_OF_RATE'] = output_df['UNIT_QTY']
    output_df['PMV_AMT'] = ""
    output_df['TOTAL_PMV'] = ""
    output_df['ACCESSORIES_FLG'] = "N"
    output_df['CESS_FLG'] = "N"
    output_df['THIRD_PARTY_FLG'] = "N"
    output_df['AR4_FLG'] = "N"
    output_df['REWARD_FLG'] = "Y"

    # Keep exact original value for TOTAL_VAL_FC
    output_df['TOTAL_VAL_FC'] = export_df2['Taxable value(in USD)']  # Exact original value

    output_df['STR_FLG'] = "N"
    output_df['END_USE'] = "GNX100"
    output_df['IGST_PAYMENT_STATUS'] = "P"
    output_df['TAXABLE_VALUE'] = pd.to_numeric(output_df["TOTAL_VAL_FC"], errors='coerce') * dollar_price
    output_df['IGST_AMOUNT'] = (pd.to_numeric(export_df2["IGST Rate (%)"], errors='coerce')) * output_df[
        "TAXABLE_VALUE"]
    output_df['SWI_STO'] = "06"
    output_df['SWI_DOO'] = "71"
    output_df['SWI_EPT'] = swi_ept_code
    output_df['SWI_UQC'] = output_df['RITC'].astype(str).str.strip().map(ritc_to_uqc).fillna("NOS")
    output_df['SWI_QTY'] = output_df['QTY_NOS']
    output_df['SWI_GCESS_AMT'] = 0
    output_df['SWI_GCESS_CUR'] = "INR"
    output_df['RODTEP_FLG'] = output_df['SCHEME_CODE'].apply(lambda x: "N" if x == "60" else "Y")
    output_df['SOURCE_STATE'] = "06"
    output_df['DBK_SRNO'] = ""
    output_df['DBK_QUANTITY'] = ""

    # Final Column Order as per your specification
    final_columns = [
        'INVOICE_SR_NO', 'ITEM_SR_NO', 'SCHEME_CODE', 'RITC', 'GOODS_DESC1', 'GOODS_DESC2', 'GOODS_DESC3',
        'QTY_NOS', 'UNIT_QTY', 'RATE_VALUE', 'NO_OF_UNIT', 'UNIT_OF_RATE', 'PMV_AMT', 'TOTAL_PMV',
        'ACCESSORIES_FLG', 'CESS_FLG', 'THIRD_PARTY_FLG', 'AR4_FLG', 'REWARD_FLG', 'TOTAL_VAL_FC',
        'STR_FLG', 'END_USE', 'IGST_PAYMENT_STATUS', 'TAXABLE_VALUE', 'IGST_AMOUNT',
        'SWI_STO', 'SWI_DOO', 'SWI_EPT', 'SWI_UQC', 'SWI_QTY', 'SWI_GCESS_AMT', 'SWI_GCESS_CUR',
        'RODTEP_FLG', 'SOURCE_STATE', 'DBK_SRNO', 'DBK_QUANTITY'
    ]

    return output_df[final_columns]


# 💾 Save to Excel buffer (for download)
def save_to_excel_buffer1(df, sheet_name="FormattedInvoice"):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    output.seek(0)
    return output


def save_to_excel_97_2003_buffer1(df, sheet_name="FormattedInvoice"):
    """
    Converts DataFrame to Excel 97-2003 (.xls) format and returns as a buffer
    """
    output = io.BytesIO()

    # Create workbook and worksheet
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet(sheet_name)

    # Write headers
    for col_idx, col_name in enumerate(df.columns):
        worksheet.write(0, col_idx, str(col_name))

    # Write data
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row):
            # Handle different types of data
            if pd.isna(value):
                worksheet.write(row_idx + 1, col_idx, '')
            elif isinstance(value, (int, float)):
                # Write numeric values as numbers
                worksheet.write(row_idx + 1, col_idx, value)
            else:
                # Write other values as strings
                worksheet.write(row_idx + 1, col_idx, str(value))

    # Save to buffer
    workbook.save(output)
    output.seek(0)
    return output


# Updated Streamlit code section for Bhajan
if selected_company == "Bhajan-sqm":
    uploaded_file = st.file_uploader("Upload Bhajan-sqm Invoice", type=["xlsx", "xls"])

    if uploaded_file:
        st.success("File Uploaded Successfully!")

        df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)

        # === Load Country Mapping ===
        try:
            country_df = pd.read_excel(COUNTRY_CODE_PATH)
            country_df['Country'] = country_df['Country'].astype(str).str.strip().str.upper()
            country_to_code = dict(zip(country_df['Country'], country_df['Code']))
            country_to_code['ANYTHING ELSE'] = 'NCPTI'
        except Exception as e:
            st.error(f"❌ Failed to load country code mapping: {e}")
            country_to_code = {}

        # === Detect Country in Raw Data ===
        detected_country = None
        for i in range(min(40, len(df_raw))):
            row = df_raw.iloc[i]
            for j in range(len(row) - 1):
                if isinstance(row[j], str) and 'country' in row[j].lower():
                    detected_country = str(row[j + 1]).strip().upper()
                    break
            if detected_country:
                break

        st.success(f"🌍 Detected Country: {detected_country if detected_country else 'NOT FOUND'}")

        # === Get SWI_EPT Code ===
        swi_ept_code = country_to_code.get(detected_country, country_to_code.get('ANYTHING ELSE', 'NCPTI'))
        swi_ept_code = st.text_input("📄 SWI_EPT Code (editable)", value=swi_ept_code)

        # --- Header Detection ---
        header_keywords = ["Sr. No.", "Product", "Description", "HSN", "Qty", "Rate", "Amount"]
        header_row_index = None

        for i, row in df_raw.iterrows():
            match_count = sum(
                any(k.lower() in str(cell).lower() for cell in row if pd.notna(cell)) for k in header_keywords)
            if match_count >= 3:
                header_row_index = i
                break

        # --- Process if Header is Found ---
        if header_row_index is not None:
            # st.info(f"Header row detected at index: {header_row_index}")
            df = pd.read_excel(uploaded_file, sheet_name=0, header=header_row_index)
            df.columns = [str(c).strip() for c in df.columns]

            # --- PRE-PROCESSING PIPELINE ---
            # 1. Remove completely blank rows
            df.dropna(how='all', inplace=True)
            df.reset_index(drop=True, inplace=True)

            trailing_keywords = ['less discount', 'discount', 'adjustment', 'deduction']

            rows_to_drop = []
            for idx, row in df.iterrows():
                row_str = ' '.join(row.dropna().astype(str)).lower()
                if any(k in row_str for k in trailing_keywords):
                    rows_to_drop.append(idx)

            if rows_to_drop:
                st.info(f"🧹 Removing {len(rows_to_drop)} trailing rows containing discount/adjustment summaries.")
                df.drop(rows_to_drop, inplace=True)
                df.reset_index(drop=True, inplace=True)

            # 2. Truncate the footer
            footer_keywords = [
                'round off', 'total invoice amount', 'in words', 'bank details',
                'bank name', 'bank a/c', 'certified that', 'authorised signatory',
                'common seal', 'total amount before tax', 'gst on reverse', 'add packing','Less Discount'
            ]
            footer_start_index = None
            for index, row in df.iterrows():
                row_as_string = ' '.join(row.dropna().astype(str)).lower()
                if any(keyword in row_as_string for keyword in footer_keywords):
                    footer_start_index = index
                    break

            if footer_start_index is not None:
                # st.info(f"Footer section detected starting at row {footer_start_index}. Truncating data...")
                df = df.iloc[:footer_start_index]
                df.reset_index(drop=True, inplace=True)

            # --- 3. Remove "Descriptive Header" Rows ---
            rows_to_drop = []
            key_value_columns = ['HSN', 'Qty.', 'Rate', 'in USD', 'Taxable value']

            for index, row in df.iterrows():
                is_sr_no_missing = pd.isna(row.get('Sr. No.'))
                existing_key_cols = [col for col in key_value_columns if col in df.columns]
                are_values_missing = row[existing_key_cols].isnull().all()

                if is_sr_no_missing and are_values_missing:
                    rows_to_drop.append(index)

            if rows_to_drop:
                st.info(f"Found and removed {len(rows_to_drop)} descriptive header rows that were not actual items.")
                df.drop(rows_to_drop, inplace=True)
                df.reset_index(drop=True, inplace=True)

            st.subheader("Detected Invoice Table (After All Cleaning)")
            st.dataframe(df)

            # --- Add option to simplify product descriptions ---
            simplify_descriptions = st.checkbox("Keep only main product name (remove text after first '/')", value=True)

            # --- CALL THE EXPORT FORMAT FUNCTION WITH SIMPLIFIED DESCRIPTIONS ---
            st.subheader("✨ Export Format(Consolidated Descriptions)")
            with st.spinner("Converting to export format..."):
                export_df2 = process_bhajan_to_export_format_alternative1(df, simplify_descriptions)
            st.dataframe(export_df2)

            st.success("✅ Export format conversion complete!")

            chosen_df = export_df2

            # Add download button for the export format
            st.subheader("📦 Final Mapped Excel Format")
            dollar_price = st.number_input("💵 Enter Dollar Price (Exchange Rate)", min_value=0.0, value=83.0, step=0.1)
            if st.button("🔄 Generate Final Excel Table"):
                final_mapped_df = create_final_mapped_excel1(
                    export_df2,
                    dollar_price,
                    uqc_mapping_path="/Users/piyanshu/PycharmProjects/pdftoexcel/SQC_1.xlsx",
                    swi_ept_code=swi_ept_code
                )
                st.session_state["final_df"] = final_mapped_df

                # Convert to Excel 97-2003 (.xls) format and store in session
                excel_buffer = save_to_excel_97_2003_buffer(final_mapped_df, sheet_name="FormattedInvoice")
                st.session_state["excel_output"] = excel_buffer

                st.success("✅ Final Excel 97-2003 Format ready!")
                st.dataframe(final_mapped_df)

            # === SAVE TO FOLDER ===
            if "excel_output" in st.session_state and st.button("📁 Add to Folder"):
                SAVE_DIRECTORY = "/Users/piyanshu/PycharmProjects/pdftoexcel/final_exports"
                os.makedirs(SAVE_DIRECTORY, exist_ok=True)

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_path = os.path.join(SAVE_DIRECTORY,
                                        f"formatted_invoice_bhajan_{timestamp}.xls")  # .xls extension for 97-2003 format
                try:
                    with open(file_path, "wb") as f:
                        f.write(st.session_state["excel_output"].getbuffer())
                    st.success(f"✅ File saved to: {file_path}")
                except Exception as e:
                    st.error(f"❌ Failed to save file: {e}")

            # === DOWNLOAD BUTTON ===
            if "excel_output" in st.session_state:
                st.download_button(
                    label="📥 Download Final Excel Table (97-2003 Format)",
                    data=st.session_state["excel_output"],
                    file_name="bhajana_invoice.xls",  # .xls extension for 97-2003 format
                    mime="application/vnd.ms-excel"  # MIME type for .xls
                )
        else:
            st.error("Could not detect a standard header row.")# bhajana ends


##JAVI HOME DECOR

def save_to_excel_97_2003_buffer(df, sheet_name="Sheet1"):
    """
    Convert DataFrame to Excel 97-2003 format and return as BytesIO buffer
    """
    import io
    import xlwt

    output = io.BytesIO()

    # Create workbook and worksheet using xlwt (for Excel 97-2003 format)
    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet(sheet_name)

    # Write headers
    for col_idx, col_name in enumerate(df.columns):
        worksheet.write(0, col_idx, str(col_name))

    # Write data
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row):
            # Handle different types of data
            if pd.isna(value):
                worksheet.write(row_idx + 1, col_idx, '')
            else:
                worksheet.write(row_idx + 1, col_idx, value)

    # Save to buffer
    workbook.save(output)
    output.seek(0)

    return output



if selected_company == "Javi":
    uploaded_file = st.file_uploader("Upload GST Invoice Excel File", type=["xlsx"])
    dollar_price = st.number_input("💵 Enter Dollar Price (Exchange Rate)", min_value=0.0, value=83.0, step=0.1)
    st.info(f"🧮 TAXABLE_VALUE will be taken from original invoice table")

    if uploaded_file:
        df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)

        # Extract company name and invoice number
        company_name = ""
        invoice_number = ""

        for row in df_raw[0].dropna().head(20):
            row_str = str(row).strip()
            if not company_name and any(
                    x in row_str.upper() for x in ["INTERNATIONAL", "LIMITED", "LTD", "PVT", "LLP"]):
                company_name = row_str.split(",")[0].strip()
            if "invoice no" in row_str.lower():
                try:
                    invoice_number = row_str.split("/")[-1].strip()
                except:
                    invoice_number = "XXXX"

        clean_co_name = company_name.upper().replace(" ", "_").replace(",", "")
        file_name = f"{clean_co_name}_{invoice_number}.xlsx"

        # --- Load Country Mapping ---
        try:
            country_df = pd.read_excel(COUNTRY_CODE_PATH)
            country_df['Country'] = country_df['Country'].astype(str).str.strip().str.upper()
            country_to_code = dict(zip(country_df['Country'], country_df['Code']))
            country_to_code['ANYTHING ELSE'] = 'NCPTI'
        except Exception as e:
            st.error(f"❌ Failed to load country code mapping: {e}")
            country_to_code = {}

        df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None) # Re-reading, as per original snippet structure

        # --- Extract Country ---
        detected_country = None
        for i in range(min(40, len(df_raw))):
            row = df_raw.iloc[i]
            for j in range(len(row) - 1):
                if isinstance(row[j], str) and 'country' in row[j].lower():
                    detected_country = str(row[j + 1]).strip().upper()
                    break
            if detected_country:
                break
        st.success(f"🌍 Detected Country: {detected_country if detected_country else 'NOT FOUND'}")

        # --- Get SWI_EPT Code ---
        swi_ept_code = country_to_code.get(detected_country, country_to_code.get('ANYTHING ELSE', 'NCPTI'))
        swi_ept_code = st.text_input("📄 SWI_EPT Code (editable)", value=swi_ept_code)

        # --- Header Detection ---
        header_keywords = ["SR. NO", "DESCRIPTION", "HSN", "QTY", "RATE", "AMOUNT", "TAXABLE"]
        header_row_index = None
        for i, row in df_raw.iterrows():
            match_count = sum(
                any(k.lower() in str(cell).lower() for cell in row if pd.notna(cell)) for k in header_keywords)
            if match_count >= 4:
                header_row_index = i
                break

        if header_row_index is not None:
            df = pd.read_excel(uploaded_file, sheet_name=0, header=header_row_index)
            df = df.dropna(how='all')
            df.columns = [str(c).strip() for c in df.columns]
            st.subheader("Detected Invoice Table")
            st.dataframe(df)

            # --- Dynamic Column Mapping for Javi (specifically for the column next to 'Unnamed: 2') ---
            unnamed2_col_idx = -1
            for i, col_name in enumerate(df.columns):
                # We need a robust way to find 'Unnamed: 2' specifically.
                # It's usually based on its exact name or position if it's consistently unnamed.
                # Assuming 'Unnamed: 2' is literally the column name
                if col_name == "Unnamed: 2":
                    unnamed2_col_idx = i
                    break

            column_next_to_unnamed2 = None
            if unnamed2_col_idx != -1 and (unnamed2_col_idx + 1) < len(df.columns):
                column_next_to_unnamed2 = df.columns[unnamed2_col_idx + 1]
                # st.info(f"Identified column next to 'Unnamed: 2': **'{column_next_to_unnamed2}'** (This column's content will be combined with DESCRIPTION)")
            elif unnamed2_col_idx != -1:
                st.warning("Found 'Unnamed: 2' but no column immediately next to it.")


            column_mapping = {
                "SR. NO.": "ITEM_SR_NO",
                "DESCRIPTION": "GOODS_DESC1",
                # Map the column *next to* 'Unnamed: 2' to a temporary detail column
                column_next_to_unnamed2: "GOODS_DESC1_DETAIL_TEMP",
                "HSN Code": "RITC",
                "QTY/UNIT": "QTY_UNIT",
                "QTY/SQ. MTR": "QTY_SQ_MTR",
                "RATE PER SQ. MTR. IN US$": "RATE_PER_SQ_MTR_USD",  # New temporary column
                "RATE SQ. MTR. IN US$": "RATE_PER_SQ_MTR_USD",    # New temporary column
                "RATE PER PC. IN INR": "RATE_PER_PC_INR_TEMP",     # New temporary column for "RATE PER PC IN INR"
                "TAXABLE VALUE (INR)": "TAXABLE_VALUE",
                "IGST": "IGST_RATE",
                "Unnamed: 14": "IGST_AMOUNT"
            }

            # Filter out None keys from mapping if column_next_to_unnamed2 wasn't found
            column_mapping = {k: v for k, v in column_mapping.items() if k is not None}

            # Create mapped dataframe
            mapped_df = pd.DataFrame()
            for col, new_col in column_mapping.items():
                if col in df.columns: # Check if the original column exists in df
                    mapped_df[new_col] = df[col]

            # --- Create QTY_NOS by using QTY_SQ_MTR first, then QTY_UNIT as fallback ---
            def determine_qty_nos(row):
                if "QTY_SQ_MTR" in row.index:
                    qty_sq_mtr = row.get("QTY_SQ_MTR")
                    if pd.notna(qty_sq_mtr) and str(qty_sq_mtr).strip().lower() not in ["", "0", "nan", "0.0"]:
                        return qty_sq_mtr

                if "QTY_UNIT" in row.index:
                    qty_unit = row.get("QTY_UNIT")
                    if pd.notna(qty_unit) and str(qty_unit).strip().lower() not in ["", "0", "nan", "0.0"]:
                        return qty_unit
                return None

            if "QTY_SQ_MTR" in mapped_df.columns or "QTY_UNIT" in mapped_df.columns:
                mapped_df["QTY_NOS"] = mapped_df.apply(determine_qty_nos, axis=1)
                # st.info("✅ **QTY_NOS** created: Using 'QTY/SQ. MTR' where available, 'QTY/UNIT' as fallback.")

            # --- Determine UNIT_QTY based on source of QTY_NOS immediately ---
            # This is crucial for the subsequent RATE_VALUE calculation
            def clean_qty(value):
                if isinstance(value, str):
                    value = value.replace('\xa0', '').strip()
                return pd.to_numeric(value, errors='coerce')


            def determine_unit_qty(row):
                qty_sq = clean_qty(row.get("QTY_SQ_MTR"))
                qty_unit = clean_qty(row.get("QTY_UNIT"))

                if pd.notna(qty_sq) and qty_sq > 0:
                    return "SQM"
                elif pd.notna(qty_unit) and qty_unit > 0:
                    return "PCS"
                else:
                    return None


            # Default to PCS if QTY_SQ_MTR is not available or empty

            mapped_df["UNIT_QTY_HELPER"] = mapped_df.apply(determine_unit_qty, axis=1)

            # --- Calculate RATE_VALUE based on UNIT_QTY_HELPER ---
            mapped_df["RATE_VALUE"] = None # Initialize column

            # Convert rate columns to numeric first
            for col_temp_rate in ["RATE_PER_SQ_MTR_USD", "RATE_PER_PC_INR_TEMP"]:
                if col_temp_rate in mapped_df.columns:
                    mapped_df[col_temp_rate] = (
                        mapped_df[col_temp_rate]
                        .astype(str)
                        .str.replace(",", "", regex=False)  # 👈 remove commas
                        .str.strip()
                    )
                    mapped_df[col_temp_rate] = pd.to_numeric(mapped_df[col_temp_rate], errors='coerce')


            def calculate_rate_value(row, dollar_price):
                if row["UNIT_QTY_HELPER"] == "SQM" and "RATE_PER_SQ_MTR_USD" in row.index:
                    return row["RATE_PER_SQ_MTR_USD"]
                elif row["UNIT_QTY_HELPER"] == "PCS" and "RATE_PER_PC_INR_TEMP" in row.index and dollar_price > 0:
                    rate_inr = row["RATE_PER_PC_INR_TEMP"]
                    if pd.notna(rate_inr):
                        return rate_inr / dollar_price
                return None

            mapped_df["RATE_VALUE"] = mapped_df.apply(lambda row: calculate_rate_value(row, dollar_price), axis=1)
            # st.info("✅ **RATE_VALUE** calculated: Uses 'RATE PER SQ. MTR. IN US$' for SQM, 'RATE PER PC IN INR' / dollar price for PCS.")

            # --- Clean and format numeric columns (BUT KEEP ORIGINAL TAXABLE_VALUE & IGST_AMOUNT) ---
            numeric_columns = ["QTY_NOS", "IGST_RATE"] # RATE_VALUE is now calculated separately
            for col in numeric_columns:
                if col in mapped_df.columns:
                    mapped_df[col] = mapped_df[col].astype(str).str.strip()
                    mapped_df[col] = pd.to_numeric(mapped_df[col], errors='coerce')

            # --- Clean TAXABLE_VALUE and IGST_AMOUNT (remove Rs. and commas) ---
            for col in ["TAXABLE_VALUE", "IGST_AMOUNT"]:
                if col in mapped_df.columns:
                    mapped_df[col] = mapped_df[col].astype(str).str.replace('Rs.', '', regex=False).str.replace(',', '', regex=False).str.strip()
                    mapped_df[col] = pd.to_numeric(mapped_df[col], errors='coerce')

            # --- Combine GOODS_DESC1 and GOODS_DESC1_DETAIL_TEMP into GOODS_DESC1 ---
            if "GOODS_DESC1" in mapped_df.columns:
                if "GOODS_DESC1_DETAIL_TEMP" in mapped_df.columns:
                    def combine_descriptions(row):
                        main_desc = str(row["GOODS_DESC1"]).strip() if pd.notna(row["GOODS_DESC1"]) else ""
                        detail_desc = str(row["GOODS_DESC1_DETAIL_TEMP"]).strip() if pd.notna(row["GOODS_DESC1_DETAIL_TEMP"]) else ""

                        if main_desc and main_desc.lower() != "nan" and detail_desc and detail_desc.lower() != "nan":
                            return f"{main_desc} - {detail_desc}"
                        elif main_desc and main_desc.lower() != "nan":
                            return main_desc
                        elif detail_desc and detail_desc.lower() != "nan":
                            return detail_desc
                        else:
                            return ""

                    mapped_df["GOODS_DESC1"] = mapped_df.apply(combine_descriptions, axis=1)
                    # Remove the temporary detail column after combining
                    mapped_df = mapped_df.drop("GOODS_DESC1_DETAIL_TEMP", axis=1)
                    # st.info("✅ Combined 'DESCRIPTION' and the column next to 'Unnamed: 2' into **GOODS_DESC1**.")
                else:
                    st.info("Only 'DESCRIPTION' column found, using it for **GOODS_DESC1**.")
            else:
                st.warning("Neither 'DESCRIPTION' nor an identifiable column next to 'Unnamed: 2' was found for **GOODS_DESC1**.")


            # --- Footer Cleanup ---
            footer_keywords = ["TOTAL", "NUMBER OF PACKAGES", "TOTAL GROSS WEIGHT", "TOTAL NET WEIGHT",
                               "AD CODE", "EC NO", "DRAWBACK", "Bank Name", "Bank A/C"]

            def is_footer_row(row):
                return any(any(kw.lower() in str(cell).lower() for kw in footer_keywords) for cell in row)

            mapped_df = mapped_df[~mapped_df.apply(is_footer_row, axis=1)].reset_index(drop=True)

            # --- Enhanced Filtering - Remove empty rows ---
            key_columns = ["ITEM_SR_NO", "RITC", "GOODS_DESC1", "QTY_NOS"]

            def is_empty_row(row):
                for col in key_columns:
                    if col in row.index:
                        value = row[col]
                        if pd.notna(value) and str(value).strip().lower() not in ["", "0", "nan", "0.0"]:
                            return False
                return True

            mapped_df = mapped_df[~mapped_df.apply(is_empty_row, axis=1)].reset_index(drop=True)

            for field in ["ITEM_SR_NO", "RITC"]:
                if field in mapped_df.columns:
                    mapped_df = mapped_df[
                        mapped_df[field].notna() &
                        ~(mapped_df[field].astype(str).str.strip().isin(["", "0", "nan", "0.0"]))
                        ]

            if "GOODS_DESC1" in mapped_df.columns:
                mapped_df = mapped_df[
                    mapped_df["GOODS_DESC1"].notna() &
                    (mapped_df["GOODS_DESC1"].astype(str).str.strip() != "")
                    ]

            mapped_df = mapped_df.reset_index(drop=True)

            # --- Calculate TOTAL_VAL_FC ---
            if "QTY_NOS" in mapped_df.columns and "RATE_VALUE" in mapped_df.columns:
                mapped_df["TOTAL_VAL_FC"] = (pd.to_numeric(mapped_df["QTY_NOS"], errors='coerce') *
                                             pd.to_numeric(mapped_df["RATE_VALUE"], errors='coerce'))

            st.subheader("📄 Cleaned Mapped Table")
            st.dataframe(mapped_df)

            # --- Final Format ---
            required_columns = [
                "INVOICE_SR_NO", "ITEM_SR_NO", "SCHEME_CODE", "RITC", "GOODS_DESC1", "GOODS_DESC2", "GOODS_DESC3",
                "QTY_NOS", "UNIT_QTY", "RATE_VALUE", "NO_OF_UNIT", "UNIT_OF_RATE", "PMV_AMT", "TOTAL_PMV",
                "ACCESSORIES_FLG", "CESS_FLG", "THIRD_PARTY_FLG", "AR4_FLG", "REWARD_FLG", "TOTAL_VAL_FC",
                "STR_FLG", "END_USE", "IGST_PAYMENT_STATUS", "TAXABLE_VALUE", "IGST_AMOUNT", "SWI_STO",
                "SWI_DOO", "SWI_EPT", "SWI_UQC", "SWI_QTY", "SWI_GCESS_AMT", "SWI_GCESS_CUR", "RODTEP_FLG",
                "SOURCE_STATE", "DBK_SRNO", "DBK_QUANTITY"
            ]

            output_df = pd.DataFrame(columns=required_columns)

            if not mapped_df.empty:
                for col in output_df.columns:
                    if col in mapped_df.columns:
                        output_df[col] = mapped_df[col]

                # --- Set values according to your specifications ---
                output_df["INVOICE_SR_NO"] = 1
                output_df["GOODS_DESC2"] = "SIZE:"
                output_df["GOODS_DESC3"] = ""

                # UNIT_QTY is already determined in UNIT_QTY_HELPER
                output_df["UNIT_QTY"] = mapped_df["UNIT_QTY_HELPER"] # Use the already determined UNIT_QTY_HELPER

                def determine_swi_qty(row):
                    if row["UNIT_QTY"] == "SQM":
                        return row["QTY_NOS"]
                    else:
                        return ""

                output_df["NO_OF_UNIT"] = 1
                output_df["UNIT_OF_RATE"] = output_df["UNIT_QTY"]
                output_df["PMV_AMT"] = ""
                output_df["TOTAL_PMV"] = ""
                output_df["ACCESSORIES_FLG"] = "N"
                output_df["CESS_FLG"] = "N"
                output_df["THIRD_PARTY_FLG"] = "N"
                output_df["AR4_FLG"] = "N"
                output_df["STR_FLG"] = "N"
                output_df["REWARD_FLG"] = "Y"
                output_df["END_USE"] = "GNX100"
                output_df["IGST_PAYMENT_STATUS"] = "P"
                output_df["SWI_STO"] = "06"
                output_df["SWI_DOO"] = "67"
                output_df["SWI_EPT"] = swi_ept_code
                output_df["SWI_QTY"] = output_df.apply(determine_swi_qty, axis=1)
                output_df["SWI_GCESS_AMT"] = "0"
                output_df["SWI_GCESS_CUR"] = "INR"
                output_df["SOURCE_STATE"] = "06"
                output_df["DBK_SRNO"] = ""
                output_df["DBK_QUANTITY"] = ""

                # TOTAL_VAL_FC is already calculated in mapped_df based on the new RATE_VALUE logic
                # So just transfer it
                if "TOTAL_VAL_FC" in mapped_df.columns:
                    output_df["TOTAL_VAL_FC"] = mapped_df["TOTAL_VAL_FC"]


                # SCHEME & RODTEP
                output_df["SCHEME_CODE"] = output_df["RITC"].astype(str).apply(
                    lambda x: "60" if x.startswith(("60", "61", "62", "63")) else "19")
                output_df["RODTEP_FLG"] = output_df["SCHEME_CODE"].apply(lambda x: "N" if x == "60" else "Y")

                # --- Load UQC Mapping ---
                try:
                    uqc_map_df = pd.read_excel(UQC_MAPPING_PATH)
                    uqc_map_df['RITC'] = uqc_map_df['RITC'].astype(str).str.strip()
                    ritc_to_uqc = dict(zip(uqc_map_df['RITC'], uqc_map_df['SWI_UQC']))
                    output_df['SWI_UQC'] = output_df['RITC'].astype(str).str.strip().map(ritc_to_uqc).fillna("SQM")
                except Exception as e:
                    st.error(f"⚠️ Failed to load UQC mapping: {e}")
                    output_df['SWI_UQC'] = "SQM"

                # --- Final check to ensure no empty rows in output ---
                output_df = output_df[output_df["ITEM_SR_NO"].notna() & (output_df["ITEM_SR_NO"] != "")].reset_index(drop=True)

                # --- Excel Output ---
                if not output_df.empty:
                    excel_buffer = save_to_excel_97_2003_buffer(output_df, sheet_name="FormattedInvoice")
                    st.session_state["excel_output"] = excel_buffer

                    st.success("✅ Final Excel 97-2003 Format ready!")
                    st.dataframe(output_df)

                    # --- SAVE TO FOLDER ---
                    if "excel_output" in st.session_state and st.button("📁 Add to Folder"):
                        SAVE_DIRECTORY = "/Users/piyanshu/PycharmProjects/pdftoexcel/final_exports"
                        os.makedirs(SAVE_DIRECTORY, exist_ok=True)

                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        file_path = os.path.join(SAVE_DIRECTORY,
                                                 f"formatted_invoice_javi_{timestamp}.xls")
                        try:
                            with open(file_path, "wb") as f:
                                f.write(st.session_state["excel_output"].getbuffer())
                            st.success(f"✅ File saved to: {file_path}")
                        except Exception as e:
                            st.error(f"❌ Failed to save file: {e}")

                    # --- DOWNLOAD BUTTON ---
                    if "excel_output" in st.session_state:
                        st.download_button(
                            label="📥 Download Final Excel Table (97-2003 Format)",
                            data=st.session_state["excel_output"],
                            file_name="javi_invoice.xls",
                            mime="application/vnd.ms-excel"
                        )
                else:
                    st.warning("⚠️ No valid data rows found after processing.")
            else:
                st.warning("⚠️ No valid data found in the uploaded file.")
        else:
            st.error("❌ Could not detect table headers. Please check your file format.")


##TAX INIDIA
# Import necessary libraries at the top of your script
import streamlit as st
import pandas as pd
import os
from datetime import datetime
import xlwt

# Create dummy files for the demo if they don't exist
if 'COUNTRY_CODE_PATH' not in globals():
    COUNTRY_CODE_PATH = "country_codes.xlsx"
if 'UQC_MAPPING_PATH' not in globals():
    UQC_MAPPING_PATH = "uqc_mapping.xlsx"

if not os.path.exists(COUNTRY_CODE_PATH):
    pd.DataFrame({'Country': ['DENMARK', 'GERMANY'], 'Code': ['DK', 'DE']}).to_excel(COUNTRY_CODE_PATH, index=False)
if not os.path.exists(UQC_MAPPING_PATH):
    pd.DataFrame({'RITC': ['61091000', '62052000'], 'SWI_UQC': ['PCS', 'PCS']}).to_excel(UQC_MAPPING_PATH, index=False)
# ----------------------------------------------------

# --- Assume `selected_company` is set from a dropdown in your Streamlit app ---
# For demonstration, we'll set it here:
# selected_company = "Tax India"

if selected_company == "Tax India":
    uploaded_file = st.file_uploader("Upload GST Invoice Excel File", type=["xlsx"])
    dollar_price = st.number_input("💵 Enter Dollar Price (Exchange Rate)", min_value=0.0, value=85.05, step=0.1)
    st.info(f"🧮 TAXABLE_VALUE will be taken from original invoice table")

    if uploaded_file:
        df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)

        # Extract company name and invoice number for Tax India
        company_name = ""
        invoice_number = ""

        # Look for company name in first few rows
        for i in range(min(10, len(df_raw))):
            for j in range(min(5, len(df_raw.columns))):
                row_str = str(df_raw.iloc[i, j]).strip() if pd.notna(df_raw.iloc[i, j]) else ""
                if "TEX INDIA" in row_str.upper() or "TAX INDIA" in row_str.upper():
                    company_name = row_str
                    break
            if company_name:
                break

        # Look for invoice number pattern (e.g., TI-148 DT. 16.06.2025)
        for i in range(min(15, len(df_raw))):
            for j in range(len(df_raw.columns)):
                cell_value = str(df_raw.iloc[i, j]).strip() if pd.notna(df_raw.iloc[i, j]) else ""
                if ("TI-" in cell_value or "T.I-" in cell_value) and "DT." in cell_value:
                    try:
                        invoice_number = cell_value.split("DT.")[0].strip().replace("TI-", "").replace("T.I-", "")
                    except Exception:
                        invoice_number = "XXXX"
                    break
            if invoice_number:
                break

        if not invoice_number:
            invoice_number = "XXXX"

        # --- Load Country Mapping ---
        try:
            country_df = pd.read_excel(COUNTRY_CODE_PATH)
            country_df['Country'] = country_df['Country'].astype(str).str.strip().str.upper()
            country_to_code = dict(zip(country_df['Country'], country_df['Code']))
            country_to_code['ANYTHING ELSE'] = 'NCPTI'
        except Exception as e:
            st.error(f"❌ Failed to load country code mapping: {e}")
            country_to_code = {}

        # --- Extract Country ---
        detected_country = None
        for i in range(min(25, len(df_raw))):
            for j in range(len(df_raw.columns)):
                cell_value = str(df_raw.iloc[i, j]).strip().upper() if pd.notna(df_raw.iloc[i, j]) else ""
                if cell_value in country_to_code:
                    detected_country = cell_value
                    break
            if detected_country:
                break

        st.success(f"🌍 Detected Country: {detected_country if detected_country else 'NOT FOUND'}")

        # --- Get SWI_EPT Code ---
        swi_ept_code = country_to_code.get(detected_country, country_to_code.get('ANYTHING ELSE', 'NCPTI'))
        swi_ept_code = st.text_input("📄 SWI_EPT Code (editable)", value=swi_ept_code)

        # --- Find the data table header row ---
        header_row_index = None
        for i in range(len(df_raw)):
            row = df_raw.iloc[i]
            if any(str(cell).strip().lower() == "sr. no." for cell in row if pd.notna(cell)):
                header_row_index = i
                break

        if header_row_index is not None:
            df = pd.read_excel(uploaded_file, sheet_name=0, header=header_row_index)
            df = df.dropna(how='all')
            df.columns = [str(c).strip() for c in df.columns]

            df = df[pd.to_numeric(df.get('Sr. No.'), errors='coerce').notna()]
            st.subheader("Detected Invoice Table")
            st.dataframe(df)

            # --- DYNAMIC Column Mapping ---
            column_mapping = {
                "ITEM_SR_NO": ["Sr. No."],
                "SUPPLIER_ARTICLE_NO": ["SUPPLIER ARTICLE NO.", "STYLE NO."],
                "BATCH_ID": ["BATCH ID"],
                "GOODS_DESC1": ["DESCRIPTION OF GOODS", "DESCRIPTION"],
                "RITC": ["HSN\nCode", "HSN Code", "HSN"],
                "COLOR": ["COLOR"],
                "QTY_NOS": ["QTY.                PCS.", "QTY.", "QTY", "QTY. PCS."],
                "SIZE_CM": ["SIZE IN C.M.", "SIZE IN CM"],
                "RATE_VALUE": ["RATE ( USD )", "RATE(USD)", "RATE (USD)"],
                "TOTAL_VAL_FC": ["TOTAL VALUE ( USD )", "TOTAL VALUE(USD)", "TOTAL VALUE (USD)"],
                "EXCHANGE_RATE": ["EXCHANGE RATE"],
                "TAXABLE_VALUE": ["TAXABLE VALUE ( INR )", "TAXABLE VALUE(INR)", "TAXABLE VALUE (INR)"],
                "IGST_RATE": ["RATE", "IGST"],
                "IGST_AMOUNT": ["AMOUNT", "TOTAL AMOUNT ( INR )", "TOTAL AMOUNT(INR)"]
            }

            mapped_df = pd.DataFrame()
            for new_col, possible_cols in column_mapping.items():
                for p_col in possible_cols:
                    clean_p_col = ' '.join(p_col.split()).lower()
                    for df_col in df.columns:
                        clean_df_col = ' '.join(str(df_col).split()).lower()
                        if clean_p_col == clean_df_col:
                            mapped_df[new_col] = df[df_col]
                            break
                    if new_col in mapped_df:
                        break

            # --- Clean and format numeric columns ---
            numeric_columns = ["QTY_NOS", "RATE_VALUE", "TOTAL_VAL_FC", "EXCHANGE_RATE", "TAXABLE_VALUE", "IGST_RATE",
                               "IGST_AMOUNT"]
            for col in numeric_columns:
                if col in mapped_df.columns:
                    mapped_df[col] = pd.to_numeric(
                        mapped_df[col].astype(str).str.replace(',', '', regex=False).str.strip(), errors='coerce')


            # --- ROBUST Enhanced Description ---
            def combine_description(row):
                parts = []
                if 'GOODS_DESC1' in row.index and pd.notna(row['GOODS_DESC1']):
                    parts.append(str(row["GOODS_DESC1"]).strip())
                if 'COLOR' in row.index and pd.notna(row['COLOR']):
                    parts.append(f"COLOR: {str(row['COLOR']).strip()}")
                if 'SIZE_CM' in row.index and pd.notna(row['SIZE_CM']):
                    parts.append(f"SIZE: {str(row['SIZE_CM']).strip()}")
                return " | ".join(parts)


            if "GOODS_DESC1" in mapped_df.columns:
                mapped_df["GOODS_DESC1"] = mapped_df.apply(combine_description, axis=1)


            # --- Validate and clean rows ---
            def is_valid_row(row):
                if 'ITEM_SR_NO' not in row.index or pd.isna(
                    pd.to_numeric(row['ITEM_SR_NO'], errors='coerce')): return False
                if 'RITC' not in row.index or pd.isna(row['RITC']): return False
                if 'QTY_NOS' not in row.index or pd.isna(row['QTY_NOS']) or row['QTY_NOS'] == 0: return False
                return True


            if not mapped_df.empty:
                mapped_df = mapped_df[mapped_df.apply(is_valid_row, axis=1)].reset_index(drop=True)

            st.subheader("📄 Cleaned Mapped Table")
            st.dataframe(mapped_df)

            # --- Final Format ---
            required_columns = [
                "INVOICE_SR_NO", "ITEM_SR_NO", "SCHEME_CODE", "RITC", "GOODS_DESC1", "GOODS_DESC2", "GOODS_DESC3",
                "QTY_NOS", "UNIT_QTY", "RATE_VALUE", "NO_OF_UNIT", "UNIT_OF_RATE", "PMV_AMT", "TOTAL_PMV",
                "ACCESSORIES_FLG", "CESS_FLG", "THIRD_PARTY_FLG", "AR4_FLG", "REWARD_FLG", "TOTAL_VAL_FC",
                "STR_FLG", "END_USE", "IGST_PAYMENT_STATUS", "TAXABLE_VALUE", "IGST_AMOUNT", "SWI_STO",
                "SWI_DOO", "SWI_EPT", "SWI_UQC", "SWI_QTY", "SWI_GCESS_AMT", "SWI_GCESS_CUR", "RODTEP_FLG",
                "SOURCE_STATE", "DBK_SRNO", "DBK_QUANTITY"
            ]

            output_df = pd.DataFrame()

            if not mapped_df.empty:
                for col in mapped_df.columns:
                    if col in required_columns:
                        output_df[col] = mapped_df[col]

                # --- Set Tax India specific values ---
                output_df["INVOICE_SR_NO"] = 1

                # --- ROBUST setting of optional description fields ---
                # output_df["GOODS_DESC3"] = "COLOR: " + mapped_df["COLOR"].fillna('').astype(
                #     str) if 'COLOR' in mapped_df.columns else ""
                output_df["GOODS_DESC3"] = ""
                output_df["GOODS_DESC2"] = "SIZE: " + mapped_df["SIZE_CM"].fillna('').astype(
                    str) if 'SIZE_CM' in mapped_df.columns else ""

                output_df["UNIT_QTY"] = "PCS"
                output_df["NO_OF_UNIT"] = 1
                output_df["UNIT_OF_RATE"] = "PCS"
                output_df["PMV_AMT"] = ""
                output_df["TOTAL_PMV"] = ""
                output_df["ACCESSORIES_FLG"] = "N"
                output_df["CESS_FLG"] = "N"
                output_df["THIRD_PARTY_FLG"] = "N"
                output_df["AR4_FLG"] = "N"
                output_df["REWARD_FLG"] = "Y"
                output_df["STR_FLG"] = "N"
                output_df["END_USE"] = "GNX100"
                output_df["IGST_PAYMENT_STATUS"] = "P"
                output_df["SWI_STO"] = "06"
                output_df["SWI_DOO"] = "71"
                output_df["SWI_EPT"] = swi_ept_code
                output_df["SWI_QTY"] = ""
                output_df["SWI_GCESS_AMT"] = "0"
                output_df["SWI_GCESS_CUR"] = "INR"
                output_df["SOURCE_STATE"] = "06"
                output_df["DBK_SRNO"] = ""
                output_df["DBK_QUANTITY"] = ""

                # SCHEME & RODTEP based on RITC
                output_df["SCHEME_CODE"] = output_df["RITC"].astype(str).apply(
                    lambda x: "60" if x.strip().startswith(("60", "61", "62", "63")) else "19")
                output_df["RODTEP_FLG"] = output_df["SCHEME_CODE"].apply(lambda x: "N" if x == "60" else "Y")

                # --- Load UQC Mapping ---
                try:
                    uqc_map_df = pd.read_excel(UQC_MAPPING_PATH)
                    uqc_map_df['RITC'] = uqc_map_df['RITC'].astype(str).str.strip()
                    ritc_to_uqc = dict(zip(uqc_map_df['RITC'], uqc_map_df['SWI_UQC']))
                    output_df['SWI_UQC'] = output_df['RITC'].astype(str).str.strip().map(ritc_to_uqc).fillna("PCS")
                except Exception as e:
                    st.error(f"⚠️ Failed to load UQC mapping: {e}")
                    output_df['SWI_UQC'] = "PCS"

                # --- Reorder columns to final required format ---
                output_df = output_df.reindex(columns=required_columns)

                # --- Final validation and Excel Output ---
                if not output_df.empty:
                    excel_buffer = save_to_excel_97_2003_buffer(output_df, sheet_name="FormattedInvoice")
                    st.session_state["excel_output"] = excel_buffer

                    st.success("✅ Final Excel 97-2003 Format ready!")
                    st.dataframe(output_df)

                    # --- SAVE TO FOLDER ---
                    if "excel_output" in st.session_state and st.button("📁 Add to Folder"):
                        SAVE_DIRECTORY = os.path.join(os.path.expanduser("~"), "Downloads", "final_exports")
                        os.makedirs(SAVE_DIRECTORY, exist_ok=True)
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        file_path = os.path.join(SAVE_DIRECTORY,
                                                 f"formatted_invoice_tax_india_{invoice_number}_{timestamp}.xls")
                        try:
                            with open(file_path, "wb") as f:
                                f.write(st.session_state["excel_output"].getbuffer())
                            st.success(f"✅ File saved to: {file_path}")
                        except Exception as e:
                            st.error(f"❌ Failed to save file: {e}")

                    # --- DOWNLOAD BUTTON ---
                    if "excel_output" in st.session_state:
                        st.download_button(
                            label="📥 Download Final Excel Table (97-2003 Format)",
                            data=st.session_state["excel_output"],
                            file_name=f"formatted_invoice_tax_india_{invoice_number}.xls",
                            mime="application/vnd.ms-excel"
                        )
                else:
                    st.warning("⚠️ No valid data rows found after processing.")
            else:
                st.warning("⚠️ Mapped dataframe is empty. No valid data could be extracted.")
        else:
            st.error("❌ Could not detect table headers ('Sr. No.'). Please check your file format.")



import pandas as pd
import streamlit as st
import re
import os
from datetime import datetime

# Assuming you have functions: save_to_excel_97_2003_buffer() and paths: COUNTRY_CODE_PATH, UQC_MAPPING_PATH

# --- NEW Processor for the SECOND Tax India format ---
import streamlit as st
import xlwt
import pandas as pd
from decimal import Decimal, getcontext, ROUND_HALF_UP

getcontext().prec = 10


def precise_mul1(qty, rate):
    try:
        qty_dec = Decimal(str(qty))
        rate_dec = Decimal(str(rate))
        result = qty_dec * rate_dec
        return result.quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)
    except:
        return Decimal('0.0000')


def extract_main_table_from_imperial_invoice1(raw_df):
    """
    Extracts the main item table from this specific invoice format.
    Handles product descriptions followed by multiple quantity rows with variants.
    """
    print("Debug: Starting extraction for this invoice format...")
    df_str = raw_df.fillna('').astype(str)

    # Find the start of the main data table
    data_start_idx = None
    for idx, row in df_str.iterrows():
        row_text = ' '.join(row.values).upper()
        # Look for the first product entry or HSN code pattern
        if any(keyword in row_text for keyword in ['ETHYLENE', 'COTTON', 'POLYESTER']) and '63059900' in row_text:
            data_start_idx = idx
            break

    if data_start_idx is None:
        # Alternative search - look for HSN codes
        for idx, row in df_str.iterrows():
            row_values = [str(cell).strip() for cell in row.values if str(cell).strip() and str(cell).strip() != 'nan']
            if any(cell.isdigit() and len(cell) == 8 for cell in row_values):
                data_start_idx = idx
                break

    if data_start_idx is None:
        st.error("Could not find the start of data table in the invoice")
        return pd.DataFrame()

    expected_columns = [
        'S_NO', 'DESCRIPTION', 'HSN_CODE', 'QTY', 'UNIT', 'RATE_PER_UNIT',
        'AMOUNT', 'TAXABLE_VALUE', 'IGST_RATE', 'IGST_AMOUNT', 'TOTAL_VALUE'
    ]

    data_rows = []
    current_description = ""
    serial_no = 1

    # Process rows from data start
    for idx in range(data_start_idx, len(df_str)):
        row = df_str.iloc[idx]
        row_values = [str(cell).strip() for cell in row.values if str(cell).strip() and str(cell).strip() != 'nan']

        if not row_values:
            continue

        row_text = ' '.join(row_values).upper()

        # Check for footer and stop
        footer_keywords = ['TOTAL', 'GRAND TOTAL', 'AMOUNT IN WORDS', 'DECLARATION', 'TERMS']
        if any(keyword in row_text for keyword in footer_keywords) and len(
                [v for v in row_values if v.replace('.', '').replace(',', '').isdigit()]) > 3:
            break

        # Check if this is a main product description row
        has_hsn_code = any(cell.isdigit() and len(cell) == 8 for cell in row_values)
        has_product_keywords = any(
            keyword in row_text for keyword in ['ETHYLENE', 'COTTON', 'POLYESTER', 'THREAD', 'FABRIC'])

        if has_product_keywords and has_hsn_code:
            # This is a main product description with HSN code
            # Extract the description (usually the longer text)
            description_parts = []
            hsn_code = ""

            for cell in row_values:
                if cell.isdigit() and len(cell) == 8:
                    hsn_code = cell
                elif len(cell) > 3 and not cell.replace('.', '').replace(',', '').isdigit():
                    description_parts.append(cell)

            current_description = ' '.join(description_parts)

            # Extract numeric values for this row - SKIP BUYER REF AND BE MORE SELECTIVE
            numeric_values = []
            for cell in row_values:
                if cell != hsn_code and cell not in description_parts:
                    cleaned_cell = cell.replace(',', '').replace('₹', '').replace('%', '').strip()
                    try:
                        num_val = float(cleaned_cell)
                        # Skip buyer reference numbers (typically large integers without decimals)
                        # Skip values that are likely buyer refs (5+ digits without decimals)
                        if '.' in cleaned_cell or num_val < 10000:  # Quantities/rates usually have decimals or are small
                            numeric_values.append(cleaned_cell)
                        # Skip large integers that are likely buyer refs
                        elif num_val >= 10000 and '.' not in cleaned_cell:
                            continue  # Skip buyer reference numbers
                        else:
                            numeric_values.append(cleaned_cell)
                    except ValueError:
                        continue

            # For main description rows, we often don't have complete quantity data
            # So we'll create a placeholder and let the variant rows fill in the details
            if current_description and hsn_code:
                # Check if we have enough numeric values for a complete row
                if len(numeric_values) >= 6:  # Complete row with qty, rate, amount, taxable, igst_rate, igst_amount
                    qty = numeric_values[0]
                    rate_per_unit = numeric_values[1]
                    amount = numeric_values[2]
                    taxable_value = numeric_values[3]
                    igst_rate = numeric_values[4]
                    igst_amount = numeric_values[5]
                    total_value = numeric_values[6] if len(numeric_values) > 6 else ""

                    data_row = [
                        serial_no,
                        current_description,
                        hsn_code,
                        qty,
                        "MTRS",
                        rate_per_unit,
                        amount,
                        taxable_value,
                        igst_rate,
                        igst_amount,
                        total_value
                    ]

                    data_rows.append(data_row)
                    serial_no += 1

                    # Debug print for main description row
                    print(f"Main Row {serial_no - 1}: QTY={qty}, Rate={rate_per_unit}, Amount={amount}, HSN={hsn_code}")
                    print(f"All numeric values found: {numeric_values}")
                # If not enough data, skip this row and wait for variant rows with complete data

        elif current_description and not has_product_keywords:
            # This might be a variant/color row under the current product
            # Check if it has quantity data
            numeric_values = []
            for cell in row_values:
                cleaned_cell = cell.replace(',', '').replace('₹', '').replace('%', '').strip()
                try:
                    num_val = float(cleaned_cell)
                    # Skip values that are likely buyer refs (5+ digits without decimals)
                    if '.' in cleaned_cell or num_val < 100000:  # Quantities/rates usually have decimals or are smaller
                        numeric_values.append(cleaned_cell)
                    else:
                        numeric_values.append(cleaned_cell)
                except ValueError:
                    continue

            # If we have enough numeric values, this is likely a data row
            if len(numeric_values) >= 3:
                # Try to find HSN code in this row or use the previous one
                hsn_code = ""
                for cell in row_values:
                    if cell.isdigit() and len(cell) == 8:
                        hsn_code = cell
                        break

                if not hsn_code:
                    # Use the HSN from previous product if available
                    if data_rows:
                        hsn_code = data_rows[-1][2]  # Previous HSN code

                # Create variant description
                variant_parts = [cell for cell in row_values if
                                 not cell.replace('.', '').replace(',', '').replace('%', '').isdigit() and len(
                                     cell) > 1]
                variant_desc = f"{current_description} - {' '.join(variant_parts)}" if variant_parts else current_description

                qty = numeric_values[0] if len(numeric_values) > 0 else ""
                unit = "MTRS"
                rate_per_unit = numeric_values[1] if len(numeric_values) > 1 else ""
                amount = numeric_values[2] if len(numeric_values) > 2 else ""
                taxable_value = numeric_values[3] if len(numeric_values) > 3 else ""
                igst_rate = numeric_values[4] if len(numeric_values) > 4 else ""
                igst_amount = numeric_values[5] if len(numeric_values) > 5 else ""
                total_value = numeric_values[6] if len(numeric_values) > 6 else ""

                data_row = [
                    serial_no,
                    variant_desc,
                    hsn_code,
                    qty,
                    unit,
                    rate_per_unit,
                    amount,
                    taxable_value,
                    igst_rate,
                    igst_amount,
                    total_value
                ]

                data_rows.append(data_row)
                serial_no += 1

    if data_rows:
        df = pd.DataFrame(data_rows, columns=expected_columns)
        return df
    else:
        return pd.DataFrame()

def clean_numeric_column1(col):
    """Clean numeric columns by removing currency symbols and commas"""
    if col.dtype == 'object':
        cleaned = col.astype(str).str.replace('₹', '').str.replace(',', '').str.strip()
        return pd.to_numeric(cleaned, errors='coerce')
    return pd.to_numeric(col, errors='coerce')


def process_imperial_to_export_format1(structured_df, dollar_price=83.0):
    """
    Process cleaned structured data to export format.
    """
    if structured_df.empty:
        return pd.DataFrame()

    df = structured_df.copy()

    # Map columns to export format
    export_df = pd.DataFrame()
    export_df["Sr. No."] = df["S_NO"]
    export_df["Product"] = df["DESCRIPTION"]
    export_df["HSN Code"] = df["HSN_CODE"]
    export_df["Qty (in pcs)"] = clean_numeric_column1(df["QTY"])
    export_df["Unit_qty"] = df["UNIT"]
    export_df["Exchange Rate (in usd/  Pcs)"] = clean_numeric_column1(df["RATE_PER_UNIT"])
    export_df["in USD"] = clean_numeric_column1(df["AMOUNT"])
    export_df["Taxable value(in USD)"] = clean_numeric_column1(df["TAXABLE_VALUE"])
    export_df["IGST Rate (%)"] = clean_numeric_column1(df["IGST_RATE"])
    export_df["IGST Amount (in Rs)"] = clean_numeric_column1(df["IGST_AMOUNT"])
    export_df["Total (in Rs)"] = clean_numeric_column1(df["TOTAL_VALUE"])

    # Clean the dataframe
    export_df = export_df.dropna(subset=["HSN Code", "Product"])

    return export_df

def create_final_mapped_excel_imperial1(export_df2, dollar_price, uqc_mapping_path=None, swi_ept_code="NCPTI"):
    """
    Generates the final mapped Excel DataFrame for Imperial invoices,
    ensuring numeric calculations and field mapping.
    """
    ritc_to_uqc = {}

    if uqc_mapping_path:
        try:
            uqc_map_df = pd.read_excel(uqc_mapping_path)
            uqc_map_df['RITC'] = uqc_map_df['RITC'].astype(str).str.strip()
            ritc_to_uqc = dict(zip(uqc_map_df['RITC'], uqc_map_df['SWI_UQC']))
        except Exception as e:
            st.error(f"⚠️ Failed to load UQC mapping: {e}")
            ritc_to_uqc = {}

    output_df = pd.DataFrame()
    output_df['INVOICE_SR_NO'] = [1] * len(export_df2)
    output_df['ITEM_SR_NO'] = export_df2['Sr. No.']
    output_df['SCHEME_CODE'] = export_df2['HSN Code'].astype(str).apply(
        lambda x: "60" if x.startswith(("60", "61", "62", "63")) else "19")
    output_df['RITC'] = export_df2['HSN Code']
    output_df['GOODS_DESC1'] = export_df2['Product']
    output_df['GOODS_DESC2'] = "SIZE:"
    output_df['GOODS_DESC3'] = "As Per Invoice"
    output_df['QTY_NOS'] = export_df2['Qty (in pcs)']
    output_df['UNIT_QTY'] = export_df2['Unit_qty']
    output_df['RATE_VALUE'] = export_df2['Exchange Rate (in usd/  Pcs)']
    output_df['NO_OF_UNIT'] = 1
    output_df['UNIT_OF_RATE'] = output_df['UNIT_QTY']
    output_df['PMV_AMT'] = ""
    output_df['TOTAL_PMV'] = ""
    output_df['ACCESSORIES_FLG'] = "N"
    output_df['CESS_FLG'] = "N"
    output_df['THIRD_PARTY_FLG'] = "N"
    output_df['AR4_FLG'] = "N"
    output_df['REWARD_FLG'] = "Y"
    output_df['TOTAL_VAL_FC'] = export_df2['Taxable value(in USD)']
    output_df['STR_FLG'] = "N"
    output_df['END_USE'] = "GNX100"
    output_df['IGST_PAYMENT_STATUS'] = "P"
    output_df['TAXABLE_VALUE'] = pd.to_numeric(output_df['TOTAL_VAL_FC'], errors='coerce')

    # Fix numeric conversion and compute correct amounts
    output_df['TOTAL_VAL_FC'] = export_df2.apply(
        lambda row: precise_mul(row['Qty (in pcs)'], row['Exchange Rate (in usd/  Pcs)']), axis=1
    )

    output_df['IGST_AMOUNT'] = pd.to_numeric(export_df2['IGST Amount (in Rs)'], errors='coerce').round(2)
    output_df['SWI_STO'] = "06"
    output_df['SWI_DOO'] = "71"
    output_df['SWI_EPT'] = swi_ept_code
    output_df['SWI_UQC'] = output_df['RITC'].astype(str).str.strip().map(ritc_to_uqc).fillna("NOS")
    output_df['SWI_QTY'] = ""
    output_df['SWI_GCESS_AMT'] = 0
    output_df['SWI_GCESS_CUR'] = "INR"
    output_df['RODTEP_FLG'] = output_df['SCHEME_CODE'].apply(lambda x: "N" if x == "60" else "Y")
    output_df['SOURCE_STATE'] = "06"
    output_df['DBK_SRNO'] = ""
    output_df['DBK_QUANTITY'] = ""

    final_columns = [
        'INVOICE_SR_NO', 'ITEM_SR_NO', 'SCHEME_CODE', 'RITC', 'GOODS_DESC1', 'GOODS_DESC2', 'GOODS_DESC3',
        'QTY_NOS', 'UNIT_QTY', 'RATE_VALUE', 'NO_OF_UNIT', 'UNIT_OF_RATE', 'PMV_AMT', 'TOTAL_PMV',
        'ACCESSORIES_FLG', 'CESS_FLG', 'THIRD_PARTY_FLG', 'AR4_FLG', 'REWARD_FLG', 'TOTAL_VAL_FC',
        'STR_FLG', 'END_USE', 'IGST_PAYMENT_STATUS', 'TAXABLE_VALUE', 'IGST_AMOUNT',
        'SWI_STO', 'SWI_DOO', 'SWI_EPT', 'SWI_UQC', 'SWI_QTY', 'SWI_GCESS_AMT', 'SWI_GCESS_CUR',
        'RODTEP_FLG', 'SOURCE_STATE', 'DBK_SRNO', 'DBK_QUANTITY'
    ]

    return output_df[final_columns]

if selected_company == "Tax India - d":
    uploaded_file = st.file_uploader("📤 Upload IMPERIAL Invoice", type=["xlsx", "xls", "csv"])

    if uploaded_file:
        st.success("✅ File uploaded successfully!")

        # Read raw file
        if uploaded_file.name.endswith(".csv"):
            df_raw = pd.read_csv(uploaded_file, header=None)
        else:
            df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)

        # Load Country Mapping
        country_to_code = load_country_mapping()

        # Extract Country
        detected_country = extract_country_from_invoice_text(df_raw, list(country_to_code.keys()))
        country_code = country_to_code.get(detected_country, "NCPTI")

        st.markdown(f"### 🌍 Destination Country: `{detected_country}`")
        st.markdown(f"### 🏷️ Country Code: `{country_code}`")
        swi_ept_code = st.text_input("📄 SWI_EPT Code (editable)", value=country_code)

        st.subheader("📄 Raw Invoice Preview")
        st.dataframe(df_raw)

        # Extract and clean main table
        structured_df = extract_main_table_from_imperial_invoice1(df_raw)

        if not structured_df.empty:
            structured_df_cleaned = clean_description_and_drop_sparse_rows(structured_df, max_empty_allowed=2)

            st.subheader("🧾 Cleaned Structured Table")
            st.dataframe(structured_df_cleaned)
            dollar_price = st.number_input("💵 Enter Dollar Price (Exchange Rate)", min_value=0.0, value=83.0,
                                           step=0.1)

            # Convert to Export Format
            export_df2 = process_imperial_to_export_format1(structured_df_cleaned, dollar_price)

            if not export_df2.empty:
                st.subheader("📦 Intermediate Export Format")
                st.dataframe(export_df2)

                # Input Dollar Price

                if st.button("🔄 Generate Final Excel Table"):
                    final_mapped_df = create_final_mapped_excel_imperial(
                        export_df2,
                        dollar_price=dollar_price,
                        uqc_mapping_path="/Users/piyanshu/PycharmProjects/pdftoexcel/SQC_1.xlsx",
                        swi_ept_code=swi_ept_code
                    )
                    st.session_state["final_df"] = final_mapped_df

                    # Convert to Excel 97-2003 format using xlwt
                    import io

                    output = io.BytesIO()
                    workbook = xlwt.Workbook()
                    worksheet = workbook.add_sheet("FormattedInvoice")

                    # Write headers
                    for col_idx, col_name in enumerate(final_mapped_df.columns):
                        worksheet.write(0, col_idx, str(col_name))

                    # Write data
                    for row_idx, row in final_mapped_df.iterrows():
                        for col_idx, value in enumerate(row):
                            if pd.isna(value):
                                worksheet.write(row_idx + 1, col_idx, '')
                            else:
                                worksheet.write(row_idx + 1, col_idx, value)

                    # Save to buffer
                    workbook.save(output)
                    output.seek(0)
                    st.session_state["excel_output"] = output

                    st.success("✅ Final Excel 97-2003 Format ready!")
                    st.dataframe(final_mapped_df)

                # Save to Folder
                if "excel_output" in st.session_state and st.button("📁 Add to Folder"):
                    import os
                    from datetime import datetime

                    SAVE_DIRECTORY = "/Users/piyanshu/PycharmProjects/pdftoexcel/final_exports"
                    os.makedirs(SAVE_DIRECTORY, exist_ok=True)

                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    file_path = os.path.join(SAVE_DIRECTORY,
                                             f"formatted_invoice_imperial_{timestamp}.xls")  # Changed to .xls
                    try:
                        with open(file_path, "wb") as f:
                            f.write(st.session_state["excel_output"].getbuffer())
                        st.success(f"✅ File saved to: {file_path}")
                    except Exception as e:
                        st.error(f"❌ Failed to save file: {e}")

                # Final Download Button
                if "excel_output" in st.session_state:
                    st.download_button(
                        label="📥 Download Final Excel Table (97-2003 Format)",
                        data=st.session_state["excel_output"],
                        file_name="final_imperial_invoice.xls",  # Changed to .xls
                        mime="application/vnd.ms-excel"  # Changed MIME type for .xls
                    )
            else:
                st.warning("⚠️ No valid data found in the structured table.")
        else:
            st.warning("⚠️ Could not detect the structured table from the invoice.")







# if selected_company == "Tax India - d":
#     uploaded_file = st.file_uploader("Upload 'Tax India - D' Invoice Excel File", type=["xlsx"])
#     dollar_price = st.number_input("💵 Enter Dollar Price (Exchange Rate)", min_value=0.0, value=85.05, step=0.1)
#     st.info("🧮 TAXABLE_VALUE, EXCHANGE RATE, and IGST will be taken from the original invoice table.")
#
#     if uploaded_file:
#         df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
#
#         # --- STEP 1: DEBUGGING - Show exactly what Pandas is reading ---
#         st.subheader("🕵️‍♂️ Diagnostic View: Raw Data Read from Excel")
#         st.warning("Inspect this table. The header row you want to find must be visible here.")
#         st.dataframe(df_raw.head(30))  # Show the first 30 rows for inspection
#
#         # --- Extract Invoice Number (no change here) ---
#         invoice_number = ""
#         # ... (invoice number logic remains the same)
#
#         # --- Load Mappings (no change here) ---
#         # ... (country and SWI code logic remains the same)
#
#         # --- STEP 2: MORE ROBUST HEADER DETECTION ---
#         header_row_index = None
#         # Use a broader and more resilient set of keywords.
#         header_keywords = ['ref', 'hs code', 'quantity', 'unit price', 'exchange', 'taxable']
#
#         st.write("---")
#         st.write(f"🔎 Searching for a header row containing at least 4 of these keywords: `{header_keywords}`")
#
#         for i in range(len(df_raw)):
#             row = df_raw.iloc[i]
#             match_count = 0
#             try:
#                 # Create a single string from the row, handling potential errors
#                 row_str = ' '.join(str(cell).strip().lower() for cell in row if pd.notna(cell))
#                 for keyword in header_keywords:
#                     if keyword in row_str:
#                         match_count += 1
#             except Exception as e:
#                 # This row might have weird data, skip it
#                 continue
#
#             # If we find a high number of keywords, it's almost certainly our header
#             if match_count >= 4:
#                 header_row_index = i
#                 st.success(f"✅ Header Detected at Excel Row Index: {header_row_index + 1}")
#                 break
#
#         if header_row_index is not None:
#             # --- The rest of the logic from before, now with a correctly identified header ---
#             df = pd.read_excel(uploaded_file, sheet_name=0, header=header_row_index)
#             # Clean column names thoroughly
#             df.columns = [str(c).replace('\n', ' ').strip() for c in df.columns]
#
#             # --- PRE-PROCESSING FOR GROUPED DATA ---
#             first_col_name = df.columns[0]
#             second_col_name = df.columns[1]
#
#             # Identify group headers (PO# rows)
#             df['GROUP_DESC'] = df[first_col_name].where(df[first_col_name].astype(str).str.contains('PO#', na=False))
#             df['GROUP_DESC'].fillna(
#                 df[second_col_name].where(df[second_col_name].astype(str).str.contains('PO#', na=False)), inplace=True)
#
#             # Forward-fill the descriptions
#             df['GROUP_DESC'] = df['GROUP_DESC'].ffill()
#
#             # Forward-fill the merged columns ('REF:' and 'COLOR')
#             df[first_col_name] = df[first_col_name].where(~df[first_col_name].astype(str).str.contains('PO#', na=False))
#             df[first_col_name] = df[first_col_name].ffill()
#             if 'COLOR' in df.columns:
#                 df['COLOR'] = df['COLOR'].ffill()
#
#             # Filter out the original group header rows to keep only data
#             df = df[pd.to_numeric(df['HS CODE'], errors='coerce').notna()].reset_index(drop=True)
#
#             st.subheader("📄 Cleaned & Processed Table")
#             st.dataframe(df)
#
#             # --- Column Mapping ---
#             # Using precise names from the image, but the logic is flexible
#             column_mapping = {
#                 "REF": [first_col_name],
#                 "COLOR": ["COLOR"],
#                 "RITC": ["HS CODE"],
#                 "QTY_NOS": ["QUANTITY PCS"],
#                 "SIZE_CM": ["SIZE IN CM."],
#                 "RATE_VALUE": ["UNIT PRICE IN USD"],
#                 "TOTAL_VAL_FC": ["AMOUNT IN USD"],
#                 "EXCHANGE_RATE": ["EXCHANGE RATE"],
#                 "TAXABLE_VALUE": ["TAXABLE VALUE IN INR"],
#                 "IGST_RATE": ["IGST %"],
#                 "IGST_AMOUNT": ["AMOUNT"],
#                 "GROUP_DESC": ["GROUP_DESC"]
#             }
#
#             mapped_df = pd.DataFrame()
#             # This mapping is safer as it finds the first match for a target column
#             for new_col, possible_names in column_mapping.items():
#                 for name in possible_names:
#                     if name in df.columns:
#                         mapped_df[new_col] = df[name]
#                         break
#
#             # --- Generate ITEM_SR_NO ---
#             mapped_df.insert(0, 'ITEM_SR_NO', range(1, len(mapped_df) + 1))
#
#
#             # --- Build GOODS_DESC1 ---
#             def combine_description(row):
#                 parts = []
#                 if 'GROUP_DESC' in row and pd.notna(row['GROUP_DESC']):
#                     parts.append(str(row["GROUP_DESC"]).strip())
#                 if 'REF' in row and pd.notna(row['REF']):
#                     parts.append(f"REF: {str(row['REF']).strip()}")
#                 return " | ".join(parts)
#
#
#             mapped_df['GOODS_DESC1'] = mapped_df.apply(combine_description, axis=1)
#
#             # --- The rest of the final formatting and output logic ---
#             # This part remains the same as it depends on `mapped_df` being correct.
#             # ... (Place the final `required_columns`, `output_df` creation, ...
#             # ... and download button logic here exactly as in the previous answer) ...
#
#             # For completeness, here is the rest of the block:
#             required_columns = [
#                 "INVOICE_SR_NO", "ITEM_SR_NO", "SCHEME_CODE", "RITC", "GOODS_DESC1", "GOODS_DESC2", "GOODS_DESC3",
#                 "QTY_NOS", "UNIT_QTY", "RATE_VALUE", "NO_OF_UNIT", "UNIT_OF_RATE", "PMV_AMT", "TOTAL_PMV",
#                 "ACCESSORIES_FLG", "CESS_FLG", "THIRD_PARTY_FLG", "AR4_FLG", "REWARD_FLG", "TOTAL_VAL_FC",
#                 "STR_FLG", "END_USE", "IGST_PAYMENT_STATUS", "TAXABLE_VALUE", "IGST_AMOUNT", "SWI_STO",
#                 "SWI_DOO", "SWI_EPT", "SWI_UQC", "SWI_QTY", "SWI_GCESS_AMT", "SWI_GCESS_CUR", "RODTEP_FLG",
#                 "SOURCE_STATE", "DBK_SRNO", "DBK_QUANTITY"
#             ]
#             output_df = pd.DataFrame()
#
#             if not mapped_df.empty:
#                 for col in mapped_df.columns:
#                     if col in required_columns:
#                         output_df[col] = mapped_df[col]
#
#                 output_df["INVOICE_SR_NO"] = 1
#                 output_df["GOODS_DESC2"] = "COLOR: " + mapped_df["COLOR"].astype(
#                     str) if "COLOR" in mapped_df.columns else ""
#                 output_df["GOODS_DESC3"] = "SIZE: " + mapped_df["SIZE_CM"].astype(
#                     str) if "SIZE_CM" in mapped_df.columns else ""
#                 output_df["UNIT_QTY"] = "PCS"
#                 output_df["NO_OF_UNIT"] = 1
#                 output_df["UNIT_OF_RATE"] = "PCS"
#                 output_df["PMV_AMT"], output_df["TOTAL_PMV"] = "", ""
#                 output_df["ACCESSORIES_FLG"], output_df["CESS_FLG"], output_df["THIRD_PARTY_FLG"], output_df[
#                     "AR4_FLG"] = "N", "N", "N", "N"
#                 output_df["REWARD_FLG"], output_df["STR_FLG"] = "Y", "N"
#                 output_df["END_USE"] = "GNX100"
#                 output_df["IGST_PAYMENT_STATUS"] = "P"
#                 output_df["SWI_STO"], output_df["SWI_DOO"], output_df["SWI_EPT"] = "06", "71", swi_ept_code
#                 output_df["SWI_QTY"], output_df["SWI_GCESS_AMT"], output_df["SWI_GCESS_CUR"] = "", "0", "INR"
#                 output_df["SOURCE_STATE"], output_df["DBK_SRNO"], output_df["DBK_QUANTITY"] = "06", "", ""
#
#                 output_df["SCHEME_CODE"] = output_df["RITC"].astype(str).apply(
#                     lambda x: "60" if str(x).strip().startswith(("60", "61", "62", "63", "57")) else "19")
#                 output_df["RODTEP_FLG"] = output_df["SCHEME_CODE"].apply(lambda x: "N" if x == "60" else "Y")
#
#                 try:
#                     uqc_map_df = pd.read_excel(UQC_MAPPING_PATH)
#                     uqc_map_df['RITC'] = uqc_map_df['RITC'].astype(str).str.strip()
#                     ritc_to_uqc = dict(zip(uqc_map_df['RITC'], uqc_map_df['SWI_UQC']))
#                     output_df['SWI_UQC'] = output_df['RITC'].astype(str).str.strip().map(ritc_to_uqc).fillna("PCS")
#                 except Exception as e:
#                     st.error(f"⚠️ Failed to load UQC mapping: {e}")
#                     output_df['SWI_UQC'] = "PCS"
#
#                 output_df = output_df.reindex(columns=required_columns).fillna('')
#                 excel_buffer = save_to_excel_97_2003_buffer(output_df, sheet_name="FormattedInvoice")
#                 st.session_state["excel_output"] = excel_buffer
#
#                 st.success("✅ Final Excel 97-2003 Format ready!")
#                 st.dataframe(output_df)
#
#                 st.download_button(
#                     label="📥 Download Final Excel Table (97-2003 Format)",
#                     data=st.session_state["excel_output"],
#                     file_name=f"formatted_invoice_tax_india_D_{invoice_number}.xls",
#                     mime="application/vnd.ms-excel"
#                 )
#         else:
#             st.error(
#                 "❌ Could not detect table headers after final attempt. Please check the 'Diagnostic View' above. The row containing headers might be missing or in an unexpected format.")

# --- NEW Processor for Tax India - d1 format ---
import streamlit as st
import xlwt
import pandas as pd
from decimal import Decimal, getcontext, ROUND_HALF_UP

getcontext().prec = 10


def precise_mul1d1(qty, rate):
    try:
        qty_dec = Decimal(str(qty))
        rate_dec = Decimal(str(rate))
        result = qty_dec * rate_dec
        return result.quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)
    except:
        return Decimal('0.0000')
def clean_numeric_column1d1(col):
    """Clean numeric columns by removing currency symbols and commas"""
    if col.dtype == 'object':
        cleaned = col.astype(str).str.replace('₹', '').str.replace(',', '').str.strip()
        return pd.to_numeric(cleaned, errors='coerce')
    return pd.to_numeric(col, errors='coerce')


def extract_main_table_from_imperial_invoice1d1(raw_df):
    """
    Extracts the main item table from Imperial/TEX INDIA invoice - d1 format.
    Based on the actual header structure visible in the image.
    """
    print("Debug: Starting extraction for TEX INDIA d1 invoice...")
    df_str = raw_df.fillna('').astype(str)

    # Look for the specific header row based on the visible headers
    header_row_idx = None

    # Search for the header row containing these specific column names
    header_indicators = ['BUYER REF', 'BUYER\'S CODE', 'Description of Goods', 'FRENCH DESIGNATION', 'COLOUR',
                         'ENGLISH DESIGNATION', 'QUANTITY', 'UNIT', 'AMOUNT', 'EXCHANGE RATE']

    for idx, row in df_str.iterrows():
        row_text = ' '.join(row.values).upper()
        # Check if this row contains multiple header indicators
        matches = sum(1 for indicator in header_indicators if indicator.upper() in row_text)
        if matches >= 4:  # If at least 4 header indicators are found
            header_row_idx = idx
            print(f"Debug: Found header row at index {idx} with {matches} matching indicators")
            break

    # Alternative search - look for specific unique headers
    if header_row_idx is None:
        for idx, row in df_str.iterrows():
            row_text = ' '.join(row.values).upper()
            if 'BUYER REF' in row_text and 'DESCRIPTION OF GOODS' in row_text:
                header_row_idx = idx
                print(f"Debug: Found header row at index {idx} using alternative method")
                break

    if header_row_idx is None:
        print("Debug: Header row not found, trying manual inspection...")
        # Print rows around potential header area for manual inspection
        for i in range(min(30, len(df_str))):
            row_text = ' '.join(df_str.iloc[i].values)
            print(f"Row {i}: {row_text[:200]}...")

        # Use a reasonable default
        header_row_idx = 10  # Adjust based on your invoice structure

    print(f"Debug: Using header row index: {header_row_idx}")

    # Extract header row to understand column positions
    header_row = df_str.iloc[header_row_idx]
    print(f"Debug: Header row content: {header_row.values}")

    # Define expected columns based on the actual headers
    expected_columns = [
        'BUYER_REF', 'SN_FROM_NO', 'BUYERS_CODE', 'DESCRIPTION_OF_GOODS',
        'FRENCH_DESIGNATION', 'COLOUR1', 'ENGLISH_DESIGNATION', 'COLOUR2',
        'SIZE_IN', 'QUANTITY', 'UNIT', 'AMOUNT', 'EXCHANGE_RATE',
        'TAXABLE_VALUE_INR', 'IGST_RATE', 'IGST_AMOUNT', 'TOTAL_AMOUNT_INR'
    ]

    data_rows = []

    # Start processing from the row after header
    for idx in range(header_row_idx + 1, len(df_str)):
        row = df_str.iloc[idx]
        row_values = [str(cell).strip() for cell in row.values if str(cell).strip() and str(cell).strip() != 'nan']

        if not row_values or len(row_values) < 3:
            continue

        row_text = ' '.join(row_values).upper()
        print(f"Debug: Processing row {idx}: {row_text[:100]}...")

        # Check for footer/total rows
        footer_keywords = ['TOTAL', 'SUB TOTAL', 'GRAND TOTAL', 'AMOUNT IN WORDS', 'DECLARATION', 'THANK YOU', 'TERMS',
                           'CONDITIONS']
        if any(keyword in row_text for keyword in footer_keywords):
            print(f"Debug: Found footer section at row {idx}, stopping extraction")
            break

        # Check if this is a data row (has product description and numeric values)
        has_product_description = any(keyword in row_text for keyword in
                                      ['COTTON', 'POLYESTER', 'HANDLOOM', 'WOVEN', 'JACQUARD', 'FABRIC', 'TEXTILE'])
        has_numeric_data = len([cell for cell in row_values if cell.replace('.', '').replace(',', '').isdigit()]) >= 2

        # Check for rows with substantial content
        has_sufficient_content = len(row_values) >= 5

        if (has_product_description or has_numeric_data) and has_sufficient_content:
            print(f"Debug: Found product data row at {idx}")

            # Get all cells from this row (including empty ones to maintain column positions)
            all_cells = [str(cell).strip() for cell in row.values]

            # Map to expected columns based on position
            # Adjust these indices based on your actual column positions
            buyer_ref = all_cells[0] if len(all_cells) > 0 else ""
            sn_from_no = all_cells[1] if len(all_cells) > 1 else ""
            buyers_code = all_cells[2] if len(all_cells) > 2 else ""
            description_of_goods = all_cells[3] if len(all_cells) > 3 else ""
            french_designation = all_cells[4] if len(all_cells) > 4 else ""
            colour1 = all_cells[5] if len(all_cells) > 5 else ""
            english_designation = all_cells[6] if len(all_cells) > 6 else ""
            colour2 = all_cells[7] if len(all_cells) > 7 else ""
            size_in = all_cells[8] if len(all_cells) > 8 else ""
            quantity = all_cells[9] if len(all_cells) > 9 else ""
            unit = all_cells[10] if len(all_cells) > 10 else ""
            amount = all_cells[11] if len(all_cells) > 11 else ""
            exchange_rate = all_cells[12] if len(all_cells) > 12 else ""
            taxable_value_inr = all_cells[13] if len(all_cells) > 13 else ""
            igst_rate = all_cells[14] if len(all_cells) > 14 else ""
            igst_amount = all_cells[15] if len(all_cells) > 15 else ""
            total_amount_inr = all_cells[16] if len(all_cells) > 16 else ""

            # Clean up empty values
            def clean_value(val):
                return val if val and val != 'nan' else ""

            # Apply cleaning
            data_row = [
                clean_value(buyer_ref),
                clean_value(sn_from_no),
                clean_value(buyers_code),
                clean_value(description_of_goods),
                clean_value(french_designation),
                clean_value(colour1),
                clean_value(english_designation),
                clean_value(colour2),
                clean_value(size_in),
                clean_value(quantity),
                clean_value(unit),
                clean_value(amount),
                clean_value(exchange_rate),
                clean_value(taxable_value_inr),
                clean_value(igst_rate),
                clean_value(igst_amount),
                clean_value(total_amount_inr)
            ]

            # Only add row if it has essential data
            if (description_of_goods or english_designation) and (quantity or amount):
                print(f"Debug: Created data row: {data_row}")
                data_rows.append(data_row)

    print(f"Debug: Total data rows found: {len(data_rows)}")

    # Create DataFrame
    if data_rows:
        df = pd.DataFrame(data_rows, columns=expected_columns)

        # Add serial numbers
        df['S_NO'] = range(1, len(df) + 1)

        # Reorder columns to put S_NO first
        cols = ['S_NO'] + [col for col in df.columns if col != 'S_NO']
        df = df[cols]

        print(f"Debug: Created DataFrame with {len(df)} rows")
        print(f"Debug: DataFrame columns: {df.columns.tolist()}")
        return df
    else:
        print("Debug: No data rows found")
        return pd.DataFrame()


def process_imperial_to_export_format1d1(structured_df, dollar_price=83.0):
    """
    Process cleaned Imperial structured data to export format for d1 format.
    Updated to handle the new column structure and include TAXABLE_VALUE_INR.
    """
    if structured_df.empty:
        return pd.DataFrame()

    df = structured_df.copy()

    # Map columns to export format using the new column names
    export_df = pd.DataFrame()
    export_df["Sr. No."] = df["S_NO"]

    # Combine description fields for better product description
    product_desc_parts = []
    if 'DESCRIPTION_OF_GOODS' in df.columns:
        product_desc_parts.append(df["DESCRIPTION_OF_GOODS"].fillna(""))
    if 'ENGLISH_DESIGNATION' in df.columns:
        product_desc_parts.append(df["ENGLISH_DESIGNATION"].fillna(""))

    if product_desc_parts:
        export_df["Product"] = product_desc_parts[0]
        for i in range(1, len(product_desc_parts)):
            mask = export_df["Product"].str.strip() == ""
            export_df.loc[mask, "Product"] = product_desc_parts[i].loc[mask]
    else:
        export_df["Product"] = "TEXTILE PRODUCT"

    # Extract HSN code - you might need to adjust this based on your data
    export_df["HSN Code"] = "57050024"  # Default HSN code, adjust as needed

    export_df["Qty (in pcs)"] = clean_numeric_column1d1(df["QUANTITY"])
    export_df["Unit_qty"] = df["UNIT"].fillna("PCS")
    export_df["Exchange Rate (in usd/  Pcs)"] = clean_numeric_column1d1(df["EXCHANGE_RATE"])
    export_df["in USD"] = clean_numeric_column1d1(df["AMOUNT"])
    export_df["Taxable value(in USD)"] = clean_numeric_column1d1(df["AMOUNT"])

    # Add TAXABLE_VALUE_INR from the structured table
    export_df["TAXABLE_VALUE_INR"] = clean_numeric_column1d1(df["TAXABLE_VALUE_INR"])

    export_df["IGST Rate (%)"] = clean_numeric_column1d1(df["IGST_RATE"])
    export_df["IGST Amount (in Rs)"] = clean_numeric_column1d1(df["IGST_AMOUNT"])
    export_df["Total (in Rs)"] = clean_numeric_column1d1(df["TOTAL_AMOUNT_INR"])

    # Additional fields
    export_df["Buyer_Ref"] = df["BUYER_REF"]
    export_df["Buyers_Code"] = df["BUYERS_CODE"]
    export_df["Colour"] = df["COLOUR1"].fillna("") + " " + df["COLOUR2"].fillna("")
    export_df["Size"] = df["SIZE_IN"]
    export_df["PO_Description"] = df["DESCRIPTION_OF_GOODS"]

    # Clean the dataframe
    export_df = export_df.dropna(subset=["Product"])

    # Clean colour field
    export_df["Colour"] = export_df["Colour"].str.strip()

    return export_df

def create_final_mapped_excel_imperial1d1(export_df2, dollar_price, uqc_mapping_path=None, swi_ept_code="NCPTI"):
    """
    Generates the final mapped Excel DataFrame for Imperial invoices d1 format.
    """
    ritc_to_uqc = {}

    if uqc_mapping_path:
        try:
            uqc_map_df = pd.read_excel(uqc_mapping_path)
            uqc_map_df['RITC'] = uqc_map_df['RITC'].astype(str).str.strip()
            ritc_to_uqc = dict(zip(uqc_map_df['RITC'], uqc_map_df['SWI_UQC']))
        except Exception as e:
            st.error(f"⚠️ Failed to load UQC mapping: {e}")
            ritc_to_uqc = {}

    output_df = pd.DataFrame()
    output_df['INVOICE_SR_NO'] = [1] * len(export_df2)
    output_df['ITEM_SR_NO'] = export_df2['Sr. No.']
    output_df['SCHEME_CODE'] = export_df2['HSN Code'].astype(str).apply(
        lambda x: "60" if x.startswith(("57", "58", "60", "61", "62", "63")) else "19")
    output_df['RITC'] = export_df2['HSN Code']
    output_df['GOODS_DESC1'] = export_df2['Product']
    output_df['GOODS_DESC2'] = export_df2['Colour'] + " - " + export_df2['Size']
    output_df['GOODS_DESC3'] = "As Per Invoice"
    output_df['QTY_NOS'] = export_df2['Qty (in pcs)']
    output_df['UNIT_QTY'] = "PCS"
    output_df['RATE_VALUE'] = export_df2['Unit_qty']  # Fixed mapping
    output_df['NO_OF_UNIT'] = 1
    output_df['UNIT_OF_RATE'] = output_df['UNIT_QTY']
    output_df['PMV_AMT'] = ""
    output_df['TOTAL_PMV'] = ""
    output_df['ACCESSORIES_FLG'] = "N"
    output_df['CESS_FLG'] = "N"
    output_df['THIRD_PARTY_FLG'] = "N"
    output_df['AR4_FLG'] = "N"
    output_df['REWARD_FLG'] = "Y"

    # Map TOTAL_VAL_FC to 'in USD' from cleaned structured table
    output_df['TOTAL_VAL_FC'] = export_df2['in USD']

    output_df['STR_FLG'] = "N"
    output_df['END_USE'] = "GNX100"
    output_df['IGST_PAYMENT_STATUS'] = "P"

    # Map TAXABLE_VALUE to 'TAXABLE_VALUE_INR' from cleaned structured table
    if 'TAXABLE_VALUE_INR' in export_df2.columns:
        output_df['TAXABLE_VALUE'] = export_df2['TAXABLE_VALUE_INR']
    else:
        # Fallback to USD value if INR not available
        output_df['TAXABLE_VALUE'] = export_df2['Taxable value(in USD)']

    output_df['IGST_AMOUNT'] = pd.to_numeric(export_df2['IGST Amount (in Rs)'], errors='coerce').round(2)
    output_df['SWI_STO'] = "06"
    output_df['SWI_DOO'] = "71"
    output_df['SWI_EPT'] = swi_ept_code
    output_df['SWI_UQC'] = output_df['RITC'].astype(str).str.strip().map(ritc_to_uqc).fillna("NOS")
    output_df['SWI_QTY'] = ""
    output_df['SWI_GCESS_AMT'] = 0
    output_df['SWI_GCESS_CUR'] = "INR"
    output_df['RODTEP_FLG'] = output_df['SCHEME_CODE'].apply(lambda x: "N" if x == "60" else "Y")
    output_df['SOURCE_STATE'] = "06"
    output_df['DBK_SRNO'] = ""
    output_df['DBK_QUANTITY'] = ""

    final_columns = [
        'INVOICE_SR_NO', 'ITEM_SR_NO', 'SCHEME_CODE', 'RITC', 'GOODS_DESC1', 'GOODS_DESC2', 'GOODS_DESC3',
        'QTY_NOS', 'UNIT_QTY', 'RATE_VALUE', 'NO_OF_UNIT', 'UNIT_OF_RATE', 'PMV_AMT', 'TOTAL_PMV',
        'ACCESSORIES_FLG', 'CESS_FLG', 'THIRD_PARTY_FLG', 'AR4_FLG', 'REWARD_FLG', 'TOTAL_VAL_FC',
        'STR_FLG', 'END_USE', 'IGST_PAYMENT_STATUS', 'TAXABLE_VALUE', 'IGST_AMOUNT',
        'SWI_STO', 'SWI_DOO', 'SWI_EPT', 'SWI_UQC', 'SWI_QTY', 'SWI_GCESS_AMT', 'SWI_GCESS_CUR',
        'RODTEP_FLG', 'SOURCE_STATE', 'DBK_SRNO', 'DBK_QUANTITY'
    ]

    return output_df[final_columns]

# Streamlit interface for Tax India - d1
if selected_company == "Tax India - d1":
    uploaded_file = st.file_uploader("📤 Upload IMPERIAL Invoice (d1 format)", type=["xlsx", "xls", "csv"])

    if uploaded_file:
        st.success("✅ File uploaded successfully!")

        # Read raw file
        if uploaded_file.name.endswith(".csv"):
            df_raw = pd.read_csv(uploaded_file, header=None)
        else:
            df_raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)

        # Load Country Mapping
        country_to_code = load_country_mapping()

        # Extract Country
        detected_country = extract_country_from_invoice_text(df_raw, list(country_to_code.keys()))
        country_code = country_to_code.get(detected_country, "NCPTI")

        st.markdown(f"### 🌍 Destination Country: `{detected_country}`")
        st.markdown(f"### 🏷️ Country Code: `{country_code}`")
        swi_ept_code = st.text_input("📄 SWI_EPT Code (editable)", value=country_code)

        st.subheader("📄 Raw Invoice Preview")
        st.dataframe(df_raw)

        # Extract and clean main table
        structured_df = extract_main_table_from_imperial_invoice1d1(df_raw)

        if not structured_df.empty:
            structured_df_cleaned = clean_description_and_drop_sparse_rows(structured_df, max_empty_allowed=2)

            st.subheader("🧾 Cleaned Structured Table")
            st.dataframe(structured_df_cleaned)

            # Input Dollar Price
            dollar_price = st.number_input("💵 Enter Dollar Price (Exchange Rate)", min_value=0.0, value=83.0, step=0.1)

            # Convert to Export Format
            export_df2 = process_imperial_to_export_format1d1(structured_df_cleaned, dollar_price)

            if not export_df2.empty:
                st.subheader("📦 Intermediate Export Format")
                st.dataframe(export_df2)

                if st.button("🔄 Generate Final Excel Table"):
                    final_mapped_df = create_final_mapped_excel_imperial1d1(
                        export_df2,
                        dollar_price=dollar_price,
                        uqc_mapping_path="/Users/piyanshu/PycharmProjects/pdftoexcel/SQC_1.xlsx",
                        swi_ept_code=swi_ept_code
                    )
                    st.session_state["final_df"] = final_mapped_df

                    # Convert to Excel 97-2003 format using xlwt
                    import io

                    output = io.BytesIO()
                    workbook = xlwt.Workbook()
                    worksheet = workbook.add_sheet("FormattedInvoice")

                    # Write headers
                    for col_idx, col_name in enumerate(final_mapped_df.columns):
                        worksheet.write(0, col_idx, str(col_name))

                    # Write data
                    for row_idx, row in final_mapped_df.iterrows():
                        for col_idx, value in enumerate(row):
                            if pd.isna(value):
                                worksheet.write(row_idx + 1, col_idx, '')
                            else:
                                worksheet.write(row_idx + 1, col_idx, value)

                    # Save to buffer
                    workbook.save(output)
                    output.seek(0)
                    st.session_state["excel_output"] = output

                    st.success("✅ Final Excel 97-2003 Format ready!")
                    st.dataframe(final_mapped_df)

                # Save to Folder
                if "excel_output" in st.session_state and st.button("📁 Add to Folder"):
                    import os
                    from datetime import datetime

                    SAVE_DIRECTORY = "/Users/piyanshu/PycharmProjects/pdftoexcel/final_exports"
                    os.makedirs(SAVE_DIRECTORY, exist_ok=True)

                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    file_path = os.path.join(SAVE_DIRECTORY, f"formatted_invoice_imperial_d1_{timestamp}.xls")
                    try:
                        with open(file_path, "wb") as f:
                            f.write(st.session_state["excel_output"].getbuffer())
                        st.success(f"✅ File saved to: {file_path}")
                    except Exception as e:
                        st.error(f"❌ Failed to save file: {e}")

                # Final Download Button
                if "excel_output" in st.session_state:
                    st.download_button(
                        label="📥 Download Final Excel Table (97-2003 Format)",
                        data=st.session_state["excel_output"],
                        file_name="final_imperial_invoice_d1.xls",
                        mime="application/vnd.ms-excel"
                    )
            else:
                st.warning("⚠️ No valid data found in the structured table.")
        else:
            st.warning("⚠️ Could not detect the structured table from the invoice.")
